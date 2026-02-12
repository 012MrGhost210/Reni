import os
import pandas as pd
from pathlib import Path, PureWindowsPath
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
import urllib.parse

# ==================== НАСТРОЙКИ ====================
# Быстро меняйте параметры поиска здесь:

# СПИСОК папок для анализа (можно указать несколько)
SOURCE_DIRECTORIES = [
    r'M:\Финансовый департамент\Treasury',  # ПЕРВАЯ папка
    r'M:\Финансовый департамент',           # ВТОРАЯ папка
    # r'D:\Документы',                      # ТРЕТЬЯ папка (раскомментировать если нужно)
    # Добавьте другие папки по мере необходимости
]

# Папка для сохранения Excel файла
OUTPUT_DIRECTORY = r'\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\Test'

# Название Excel файла с отчетом
EXCEL_FILENAME = "анализ_файлов.xlsx"

# ==================== НАСТРОЙКИ КЛЮЧЕВЫХ СЛОВ ====================

# Включить/выключить фильтрацию по ключевым словам
ENABLE_KEYWORD_FILTER = True  # True - фильтровать по ключевым словам, False - все файлы

# ОТКУДА БРАТЬ КЛЮЧЕВЫЕ СЛОВА:
# Вариант 1: Из Excel файла (раскомментируйте и укажите путь)
KEYWORDS_EXCEL_FILE = r'\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\ключевые_слова.xlsx'  # ЗАМЕНИТЕ НА СВОЙ ПУТЬ
KEYWORDS_EXCEL_SHEET = 'Лист1'  # Название листа (по умолчанию 'Лист1')
KEYWORDS_EXCEL_COLUMN = 'A'     # Колонка с ключевыми словами (A, B, C и т.д.)

# Вариант 2: Из списка в коде (раскомментируйте если не используете Excel)
# KEYWORDS = [
#     "Чешенко",
#     "отчет",
#     "финанс",
#     # "2024",  # Добавьте свои ключевые слова
#     # "смета",
# ]

# ==================== НАСТРОЙКИ ПОИСКА ====================

# Чувствительность к регистру при поиске
CASE_SENSITIVE_SEARCH = False  # True - учитывает регистр, False - не учитывает

# Где искать ключевые слова
SEARCH_IN_FILENAME_ONLY = False  # True - только в имени файла, False - в полном пути

# ==================== ДОПОЛНИТЕЛЬНЫЕ НАСТРОЙКИ ====================

# Показывать подробный процесс работы
SHOW_DETAILS = True

# Создать папку для отчета, если её нет
CREATE_OUTPUT_DIR = True

# Делать ли гиперссылки на файлы
CREATE_HYPERLINKS = True

# Открывать ли Excel файл после создания
OPEN_EXCEL_AFTER_CREATION = True

# ==================== КОНЕЦ НАСТРОЕК ====================

def load_keywords_from_excel():
    """
    Загружает ключевые слова из Excel файла
    Возвращает список ключевых слов или пустой список в случае ошибки
    """
    try:
        if not KEYWORDS_EXCEL_FILE or not os.path.exists(KEYWORDS_EXCEL_FILE):
            print(f"⚠️ Excel файл с ключевыми словами не найден: {KEYWORDS_EXCEL_FILE}")
            return []
        
        print(f"📖 Загружаем ключевые слова из: {KEYWORDS_EXCEL_FILE}")
        
        # Определяем букву колонки для pandas
        col_letter = KEYWORDS_EXCEL_COLUMN.upper()
        col_index = ord(col_letter) - 65  # A=0, B=1, C=2 и т.д.
        
        # Читаем Excel файл
        df = pd.read_excel(KEYWORDS_EXCEL_FILE, sheet_name=KEYWORDS_EXCEL_SHEET, header=None)
        
        # Получаем значения из указанной колонки
        keywords = []
        for idx, row in df.iterrows():
            if col_index < len(row):
                value = row[col_index]
                # Проверяем, что значение не пустое и это строка
                if pd.notna(value) and isinstance(value, (str, int, float)):
                    keywords.append(str(value).strip())
        
        # Убираем дубликаты и пустые строки
        keywords = list(set([k for k in keywords if k]))
        
        print(f"✅ Загружено ключевых слов: {len(keywords)}")
        if SHOW_DETAILS and keywords:
            print(f"   Примеры: {', '.join(keywords[:10])}{'...' if len(keywords) > 10 else ''}")
        
        return keywords
        
    except Exception as e:
        print(f"❌ Ошибка при загрузке ключевых слов из Excel: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_keywords():
    """
    Возвращает список ключевых слов в зависимости от настроек
    """
    # Пробуем загрузить из Excel
    excel_keywords = load_keywords_from_excel()
    if excel_keywords:
        return excel_keywords
    
    # Если не получилось загрузить из Excel, используем список из кода
    try:
        if KEYWORDS and isinstance(KEYWORDS, list):
            print("📝 Используем ключевые слова из кода")
            return KEYWORDS
    except NameError:
        # Переменная KEYWORDS не определена
        pass
    
    print("⚠️ Нет ключевых слов для поиска!")
    return []

def format_excel_file(worksheet, total_rows, enable_filter, keywords_count):
    """
    Форматирует Excel файл: настраивает ширину столбцов, стили, гиперссылки
    """
    # Устанавливаем ширину столбцов
    column_widths = {
        'A': 40,   # Имя файла
        'B': 20,   # Тип файла
        'C': 25,   # Дата изменения
        'D': 100,  # Полный путь
        'E': 30,   # Источник (папка поиска)
        'F': 50,   # Найденные ключевые слова
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Форматируем заголовки
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col in range(1, 7):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Настраиваем стиль для гиперссылок
    hyperlink_font = Font(color="0563C1", underline="single")
    
    # Применяем стиль гиперссылок
    for row in range(2, total_rows + 2):
        path_cell = worksheet.cell(row=row, column=4)
        
        if CREATE_HYPERLINKS and path_cell.hyperlink:
            path_cell.font = hyperlink_font
            
            name_cell = worksheet.cell(row=row, column=1)
            name_cell.hyperlink = path_cell.hyperlink
            name_cell.font = hyperlink_font
    
    # Делаем автофильтр
    worksheet.auto_filter.ref = f"A1:F{total_rows + 1}"
    worksheet.freeze_panes = "A2"
    
    # Условное форматирование для дат
    date_column_letter = 'C'
    date_range = f"{date_column_letter}2:{date_column_letter}{total_rows + 1}"
    
    today_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    today_font = Font(color="006100")
    
    today_rule = FormulaRule(
        formula=[f'AND(${date_column_letter}2>=TODAY(), ${date_column_letter}2<TODAY()+1)'],
        fill=today_fill,
        font=today_font
    )
    worksheet.conditional_formatting.add(date_range, today_rule)
    
    old_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    old_font = Font(color="9C0006")
    
    old_rule = FormulaRule(
        formula=[f'${date_column_letter}2<TODAY()-30'],
        fill=old_fill,
        font=old_font
    )
    worksheet.conditional_formatting.add(date_range, old_rule)
    
    # Подсветка найденных файлов
    if enable_filter:
        match_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in range(2, total_rows + 2):
            keyword_cell = worksheet.cell(row=row, column=6)
            if keyword_cell.value and keyword_cell.value != "Нет совпадений" and keyword_cell.value != "Фильтр отключен":
                for col in range(1, 7):
                    worksheet.cell(row=row, column=col).fill = match_fill

def create_file_hyperlink(file_path):
    """
    Создает корректную гиперссылку для файла
    """
    try:
        if not os.path.exists(file_path):
            return None
        
        abs_path = os.path.abspath(file_path)
        hyperlink_path = abs_path.replace('/', '\\')
        
        return hyperlink_path
    except:
        return None

def check_keywords_match(file_path, filename, keywords, case_sensitive, search_in_filename_only):
    """
    Проверяет, содержит ли файл/путь указанные ключевые слова
    Возвращает список найденных ключевых слов
    """
    matched_keywords = []
    
    if not keywords or not ENABLE_KEYWORD_FILTER:
        return []
    
    search_text = filename if search_in_filename_only else str(file_path)
    
    if not case_sensitive:
        search_text = search_text.lower()
    
    for keyword in keywords:
        if not keyword.strip():
            continue
            
        search_keyword = keyword if case_sensitive else keyword.lower()
        
        if search_keyword in search_text:
            matched_keywords.append(keyword)
    
    return matched_keywords

def should_save_file(matched_keywords, enable_filter):
    """
    Определяет, нужно ли сохранять файл в отчете
    """
    if not enable_filter:
        return True
    else:
        return bool(matched_keywords)

def create_excel_report(files_data, output_path, total_processed, matching_files, 
                       enable_filter, keywords_list):
    """
    Создает Excel файл со списком файлов с гиперссылками
    """
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Все файлы" if not enable_filter else "Отфильтрованные файлы"
        
        # Заголовки
        headers = [
            "Имя файла", 
            "Тип файла", 
            "Дата изменения", 
            "Полный путь",
            "Источник (папка)",
            "Найденные ключевые слова" if enable_filter else "Примечание"
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Заполняем данные
        for row_idx, file_info in enumerate(files_data, 2):
            (filename, file_type, mod_date, full_path, 
             source_dir, matched_keywords) = file_info
            
            worksheet.cell(row=row_idx, column=1, value=filename)
            worksheet.cell(row=row_idx, column=2, value=file_type)
            
            try:
                date_obj = datetime.strptime(mod_date, '%Y-%m-%d %H:%M:%S')
                worksheet.cell(row=row_idx, column=3, value=date_obj)
                worksheet.cell(row=row_idx, column=3).number_format = 'YYYY-MM-DD HH:MM:SS'
            except:
                worksheet.cell(row=row_idx, column=3, value=mod_date)
            
            path_cell = worksheet.cell(row=row_idx, column=4, value=full_path)
            worksheet.cell(row=row_idx, column=5, value=source_dir)
            
            if enable_filter:
                if matched_keywords:
                    keywords_str = ", ".join(matched_keywords)
                    worksheet.cell(row=row_idx, column=6, value=keywords_str)
                else:
                    worksheet.cell(row=row_idx, column=6, value="Нет совпадений")
            else:
                worksheet.cell(row=row_idx, column=6, value="Фильтр отключен")
            
            if CREATE_HYPERLINKS:
                hyperlink = create_file_hyperlink(full_path)
                if hyperlink:
                    path_cell.hyperlink = hyperlink
                    name_cell = worksheet.cell(row=row_idx, column=1)
                    name_cell.hyperlink = hyperlink
        
        # Форматирование
        total_rows = len(files_data)
        format_excel_file(worksheet, total_rows, enable_filter, len(keywords_list))
        
        # Информационные строки
        info_row = total_rows + 3
        
        worksheet.cell(row=info_row, column=1, 
                      value=f"Всего обработано файлов: {total_processed}")
        worksheet.cell(row=info_row, column=1).font = Font(bold=True, size=12)
        
        info_row += 1
        if enable_filter:
            worksheet.cell(row=info_row, column=1, 
                          value=f"Соответствует фильтрам: {matching_files} ({matching_files/total_processed*100:.1f}%)")
            worksheet.cell(row=info_row, column=1).font = Font(bold=True, color="00B050", size=12)
            
            info_row += 1
            if keywords_list:
                keywords_str = ", ".join(keywords_list[:20])
                if len(keywords_list) > 20:
                    keywords_str += f" и еще {len(keywords_list) - 20} слов"
                worksheet.cell(row=info_row, column=1, 
                              value=f"Ключевые слова ({len(keywords_list)}): {keywords_str}")
            else:
                worksheet.cell(row=info_row, column=1, 
                              value="Ключевые слова: не заданы")
        else:
            worksheet.cell(row=info_row, column=1, 
                          value=f"Режим: полный анализ всех файлов (фильтр отключен)")
            worksheet.cell(row=info_row, column=1).font = Font(bold=True, color="4472C4", size=12)
        
        # Инструкция по гиперссылкам
        if CREATE_HYPERLINKS:
            info_row += 2
            worksheet.cell(row=info_row, column=1, 
                          value="💡 ИНСТРУКЦИЯ: Щелкните по имени файла или пути, чтобы открыть файл")
            worksheet.cell(row=info_row, column=1).font = Font(color="00B050", italic=True, size=11)
        
        workbook.save(output_path)
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при создании Excel отчета: {e}")
        import traceback
        traceback.print_exc()
        return False

def open_excel_file(file_path):
    try:
        os.startfile(file_path)
        return True
    except Exception as e:
        print(f"⚠️ Не удалось открыть Excel файл: {e}")
        return False

def analyze_directory_files():
    """
    Анализирует файлы в нескольких директориях с фильтрацией по ключевым словам
    """
    if CREATE_OUTPUT_DIR:
        Path(OUTPUT_DIRECTORY).mkdir(parents=True, exist_ok=True)
    
    excel_path = Path(OUTPUT_DIRECTORY) / EXCEL_FILENAME
    files_data = []
    
    # Загружаем ключевые слова
    keywords_list = []
    if ENABLE_KEYWORD_FILTER:
        keywords_list = get_keywords()
    
    # Фильтруем существующие директории
    valid_directories = []
    for dir_path in SOURCE_DIRECTORIES:
        if os.path.exists(dir_path):
            valid_directories.append(dir_path)
        else:
            print(f"⚠️ Папка не существует: {dir_path}")
    
    if not valid_directories:
        print(f"❌ Ошибка: Ни одна из указанных папок не существует!")
        return
    
    # Счетчики
    total_processed = 0
    total_files_in_dirs = 0
    matching_files = 0
    
    print("=" * 80)
    print("🔍 АНАЛИЗ ФАЙЛОВ В НЕСКОЛЬКИХ ДИРЕКТОРИЯХ")
    print("=" * 80)
    
    print(f"📁 Папки для анализа ({len(valid_directories)}):")
    for i, dir_path in enumerate(valid_directories, 1):
        print(f"   {i}. {dir_path}")
    
    print(f"\n⚙️  Настройки поиска:")
    print(f"   Фильтр по ключевым словам: {'ВКЛЮЧЕН' if ENABLE_KEYWORD_FILTER else 'ВЫКЛЮЧЕН'}")
    
    if ENABLE_KEYWORD_FILTER:
        if keywords_list:
            print(f"   Ключевые слова: {len(keywords_list)} шт.")
            print(f"   Первые 10: {', '.join(keywords_list[:10])}{'...' if len(keywords_list) > 10 else ''}")
            print(f"   Источник: Excel файл ({KEYWORDS_EXCEL_FILE})")
        else:
            print(f"   ⚠️ Ключевые слова не загружены! Фильтр включен, но слов нет.")
        print(f"   Чувствительность к регистру: {'Да' if CASE_SENSITIVE_SEARCH else 'Нет'}")
        print(f"   Искать только в именах файлов: {'Да' if SEARCH_IN_FILENAME_ONLY else 'Нет'}")
        print(f"   Режим: ТОЛЬКО файлы с ключевыми словами")
    else:
        print(f"   Режим: ВСЕ файлы")
    
    print(f"🔗 Гиперссылки: {'ВКЛЮЧЕНЫ' if CREATE_HYPERLINKS else 'ВЫКЛЮЧЕНЫ'}")
    print(f"📊 Отчет будет сохранен: {excel_path}")
    print("-" * 80)
    
    # Обрабатываем каждую директорию
    for source_dir in valid_directories:
        print(f"\n📂 Анализируем папку: {source_dir}")
        
        dir_file_count = 0
        dir_matching_count = 0
        
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                total_files_in_dirs += 1
                file_path = Path(root) / file
                
                try:
                    filename = file_path.name
                    
                    file_extension = file_path.suffix.lower()
                    if file_extension:
                        file_type = file_extension.lstrip('.').upper()
                    else:
                        file_type = "БЕЗ РАСШИРЕНИЯ"
                    
                    try:
                        mod_time = os.path.getmtime(file_path)
                        mod_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        mod_date = 'НЕДОСТУПНО'
                    
                    full_path = str(file_path.resolve())
                    
                    matched_keywords = []
                    if ENABLE_KEYWORD_FILTER and keywords_list:
                        matched_keywords = check_keywords_match(
                            file_path, filename, keywords_list, 
                            CASE_SENSITIVE_SEARCH, SEARCH_IN_FILENAME_ONLY
                        )
                    
                    should_save = should_save_file(matched_keywords, ENABLE_KEYWORD_FILTER)
                    
                    if should_save:
                        files_data.append([
                            filename,
                            file_type,
                            mod_date,
                            full_path,
                            source_dir,
                            matched_keywords
                        ])
                        
                        if matched_keywords:
                            matching_files += 1
                            dir_matching_count += 1
                    
                    total_processed += 1
                    dir_file_count += 1
                    
                    if SHOW_DETAILS and total_processed % 500 == 0:
                        print(f"   📊 Обработано файлов: {total_processed}...")
                        
                except Exception as e:
                    print(f"   ⚠️ Ошибка при обработке файла {file_path}: {e}")
                    continue
        
        print(f"   ✅ Обработано файлов в этой папке: {dir_file_count}")
        if ENABLE_KEYWORD_FILTER and keywords_list:
            print(f"   🔍 Соответствует фильтрам: {dir_matching_count}")
    
    # Создаем Excel отчет
    if files_data:
        print("\n" + "-" * 80)
        print("📈 СОЗДАНИЕ ОТЧЕТА...")
        
        success = create_excel_report(files_data, excel_path, total_processed, 
                                     matching_files, ENABLE_KEYWORD_FILTER, keywords_list)
        
        print("-" * 80)
        print("🎯 ИТОГОВЫЕ РЕЗУЛЬТАТЫ:")
        print(f"   📄 Всего файлов в папках: {total_files_in_dirs}")
        print(f"   ✅ Успешно обработано: {total_processed}")
        
        if ENABLE_KEYWORD_FILTER:
            if keywords_list:
                print(f"   🔍 Соответствует фильтрам: {matching_files} ({matching_files/total_processed*100:.1f}%)")
                print(f"   📋 Записей в Excel (только с ключевыми словами): {len(files_data)}")
            else:
                print(f"   ⚠️ Фильтр включен, но ключевые слова не загружены!")
                print(f"   📋 Записей в Excel (нет фильтрации): {len(files_data)}")
        else:
            print(f"   📋 Записей в Excel (все файлы): {len(files_data)}")
        
        if success:
            print(f"\n   ✅ Excel отчет успешно создан: {excel_path}")
            
            if CREATE_HYPERLINKS:
                print(f"   🔗 Гиперссылки добавлены к именам файлов и путям")
                print(f"   💡 В Excel: щелкните по имени файла или пути для открытия")
            
            if OPEN_EXCEL_AFTER_CREATION:
                print(f"   📂 Открываю Excel файл...")
                open_excel_file(excel_path)
        else:
            print("   ❌ Не удалось создать Excel отчет")
    else:
        print("\nℹ️  Не найдено файлов.")
        if ENABLE_KEYWORD_FILTER:
            print("   Попробуйте изменить ключевые слова или отключить фильтрацию.")
    
    print("=" * 80)

if __name__ == "__main__":
    analyze_directory_files()
