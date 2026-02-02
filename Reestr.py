import os
import pandas as pd
from pathlib import Path, PureWindowsPath
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
import urllib.parse

# ==================== НАСТРОЙКИ ====================
# Быстро меняйте параметры поиска здесь:

# СПИСОК папок для анализа (можно указать несколько)
SOURCE_DIRECTORIES = [
    r'M:\Финансовый департамент\Treasury',  # ПЕРВАЯ папка
    r'M:\Финансовый департамент',           # ВТОРАЯ папка
    # r'D:\Документы',                      # ТРЕТЬЯ папка (раскомментировать если нужно)
    # Добавьте другие папки по мере необходимости
]

# Папка для сохранения Excel файла
OUTPUT_DIRECTORY = r'\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\Test'

# Название Excel файла
EXCEL_FILENAME = "анализ_файлов.xlsx"

# ==================== НАСТРОЙКИ ПОИСКА ====================

# Включить/выключить фильтрацию по ключевым словам
ENABLE_KEYWORD_FILTER = True  # True - фильтровать по ключевым словам, False - все файлы

# Ключевые слова для поиска (список)
KEYWORDS = [
    "Чешенко",
    "отчет",
    "финанс",
    # "2024",  # Добавьте свои ключевые слова
    # "смета",
]

# Чувствительность к регистру при поиске
CASE_SENSITIVE_SEARCH = False  # True - учитывает регистр, False - не учитывает

# Где искать ключевые слова
SEARCH_IN_FILENAME_ONLY = False  # True - только в имени файла, False - в полном пути

# ==================== ДОПОЛНИТЕЛЬНЫЕ НАСТРОЙКИ ====================

# Показывать подробный процесс работы
SHOW_DETAILS = True

# Создать папку для отчета, если её нет
CREATE_OUTPUT_DIR = True

# Делать ли гиперссылки на файлы
CREATE_HYPERLINKS = True

# Открывать ли Excel файл после создания
OPEN_EXCEL_AFTER_CREATION = True

# Сохранять ли файлы, не соответствующие фильтрам (если фильтры включены)
SAVE_NON_MATCHING_FILES = True  # True - сохранять все файлы, False - только соответствующие фильтрам

# ==================== КОНЕЦ НАСТРОЕК ====================

def format_excel_file(worksheet, total_rows):
    """
    Форматирует Excel файл: настраивает ширину столбцов, стили, гиперссылки
    """
    # Устанавливаем ширину столбцов
    column_widths = {
        'A': 40,   # Имя файла
        'B': 20,   # Тип файла
        'C': 25,   # Дата изменения
        'D': 100,  # Полный путь
        'E': 30,   # Источник (папка поиска)
        'F': 50,   # Найденные ключевые слова
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Форматируем заголовки
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col in range(1, 7):  # 6 колонок
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Настраиваем стиль для гиперссылок
    hyperlink_font = Font(color="0563C1", underline="single")
    
    # Применяем стиль гиперссылок к ячейкам с путями (колонка D)
    for row in range(2, total_rows + 2):
        path_cell = worksheet.cell(row=row, column=4)  # Колонка D
        
        if CREATE_HYPERLINKS and path_cell.hyperlink:
            path_cell.font = hyperlink_font
            
            # Также делаем гиперссылку в имени файла (колонка A)
            name_cell = worksheet.cell(row=row, column=1)
            name_cell.hyperlink = path_cell.hyperlink
            name_cell.font = hyperlink_font
    
    # Делаем автофильтр для заголовков
    worksheet.auto_filter.ref = f"A1:F{total_rows + 1}"
    
    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"
    
    # Добавляем условное форматирование для дат
    date_column_letter = 'C'
    date_range = f"{date_column_letter}2:{date_column_letter}{total_rows + 1}"
    
    # Форматирование для сегодняшних файлов
    today_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    today_font = Font(color="006100")
    
    today_rule = FormulaRule(
        formula=[f'AND(${date_column_letter}2>=TODAY(), ${date_column_letter}2<TODAY()+1)'],
        fill=today_fill,
        font=today_font
    )
    worksheet.conditional_formatting.add(date_range, today_rule)
    
    # Форматирование для старых файлов (старше 30 дней)
    old_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    old_font = Font(color="9C0006")
    
    old_rule = FormulaRule(
        formula=[f'${date_column_letter}2<TODAY()-30'],
        fill=old_fill,
        font=old_font
    )
    worksheet.conditional_formatting.add(date_range, old_rule)
    
    # Форматирование для файлов, соответствующих фильтрам (если фильтры включены)
    if ENABLE_KEYWORD_FILTER and not SAVE_NON_MATCHING_FILES:
        match_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in range(2, total_rows + 2):
            keyword_cell = worksheet.cell(row=row, column=6)  # Колонка F
            if keyword_cell.value and keyword_cell.value != "Нет совпадений":
                for col in range(1, 7):
                    worksheet.cell(row=row, column=col).fill = match_fill

def create_file_hyperlink(file_path):
    """
    Создает корректную гиперссылку для файла
    """
    try:
        # Проверяем существование файла
        if not os.path.exists(file_path):
            return None
        
        # Создаем гиперссылку для Windows
        abs_path = os.path.abspath(file_path)
        hyperlink_path = abs_path.replace('/', '\\')
        
        return hyperlink_path
    except:
        return None

def check_keywords_match(file_path, filename, keywords, case_sensitive, search_in_filename_only):
    """
    Проверяет, содержит ли файл/путь указанные ключевые слова
    Возвращает список найденных ключевых слов
    """
    matched_keywords = []
    
    if not keywords or not ENABLE_KEYWORD_FILTER:
        return []
    
    # Подготавливаем строки для поиска
    search_text = filename if search_in_filename_only else str(file_path)
    
    if not case_sensitive:
        search_text = search_text.lower()
    
    for keyword in keywords:
        if not keyword.strip():
            continue
            
        search_keyword = keyword if case_sensitive else keyword.lower()
        
        if search_keyword in search_text:
            matched_keywords.append(keyword)
    
    return matched_keywords

def create_excel_report(files_data, output_path, total_processed, total_found):
    """
    Создает Excel файл со списком файлов с гиперссылками
    """
    try:
        # Создаем новую рабочую книгу
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Все файлы"
        
        # Добавляем заголовки
        headers = [
            "Имя файла", 
            "Тип файла", 
            "Дата изменения", 
            "Полный путь",
            "Источник (папка)",
            "Найденные ключевые слова"
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Заполняем данные
        for row_idx, file_info in enumerate(files_data, 2):
            (filename, file_type, mod_date, full_path, 
             source_dir, matched_keywords) = file_info
            
            # Имя файла
            worksheet.cell(row=row_idx, column=1, value=filename)
            
            # Тип файла
            worksheet.cell(row=row_idx, column=2, value=file_type)
            
            # Дата изменения
            try:
                date_obj = datetime.strptime(mod_date, '%Y-%m-%d %H:%M:%S')
                worksheet.cell(row=row_idx, column=3, value=date_obj)
                worksheet.cell(row=row_idx, column=3).number_format = 'YYYY-MM-DD HH:MM:SS'
            except:
                worksheet.cell(row=row_idx, column=3, value=mod_date)
            
            # Полный путь
            path_cell = worksheet.cell(row=row_idx, column=4, value=full_path)
            
            # Источник (папка поиска)
            worksheet.cell(row=row_idx, column=5, value=source_dir)
            
            # Найденные ключевые слова
            if matched_keywords:
                keywords_str = ", ".join(matched_keywords)
                worksheet.cell(row=row_idx, column=6, value=keywords_str)
                
                # Подсвечиваем строку если найдены ключевые слова
                if ENABLE_KEYWORD_FILTER:
                    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    for col in range(1, 7):
                        worksheet.cell(row=row_idx, column=col).fill = fill
            else:
                worksheet.cell(row=row_idx, column=6, value="Нет совпадений")
            
            # Создаем гиперссылки если нужно
            if CREATE_HYPERLINKS:
                hyperlink = create_file_hyperlink(full_path)
                if hyperlink:
                    path_cell.hyperlink = hyperlink
                    
                    # Также в имени файла
                    name_cell = worksheet.cell(row=row_idx, column=1)
                    name_cell.hyperlink = hyperlink
        
        # Применяем форматирование
        total_rows = len(files_data)
        format_excel_file(worksheet, total_rows)
        
        # Добавляем информационные строки
        info_row = total_rows + 3
        
        # Общая информация
        worksheet.cell(row=info_row, column=1, 
                      value=f"Всего обработано файлов: {total_processed}")
        worksheet.cell(row=info_row, column=1).font = Font(bold=True, size=12)
        
        info_row += 1
        if ENABLE_KEYWORD_FILTER:
            worksheet.cell(row=info_row, column=1, 
                          value=f"Соответствует фильтрам: {total_found}")
            worksheet.cell(row=info_row, column=1).font = Font(bold=True, color="00B050", size=12)
            
            info_row += 1
            keywords_str = ", ".join(KEYWORDS) if KEYWORDS else "не заданы"
            worksheet.cell(row=info_row, column=1, 
                          value=f"Ключевые слова: {keywords_str}")
        
        # Инструкция по гиперссылкам
        if CREATE_HYPERLINKS:
            info_row += 2
            worksheet.cell(row=info_row, column=1, 
                          value="💡 ИНСТРУКЦИЯ: Щелкните по имени файла или пути, чтобы открыть файл")
            worksheet.cell(row=info_row, column=1).font = Font(color="00B050", italic=True, size=11)
        
        # Сохраняем файл
        workbook.save(output_path)
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при создании Excel отчета: {e}")
        import traceback
        traceback.print_exc()
        return False

def open_excel_file(file_path):
    """
    Открывает Excel файл после создания
    """
    try:
        os.startfile(file_path)
        return True
    except Exception as e:
        print(f"⚠️ Не удалось открыть Excel файл: {e}")
        return False

def analyze_directory_files():
    """
    Анализирует файлы в нескольких директориях с фильтрацией по ключевым словам
    """
    if CREATE_OUTPUT_DIR:
        Path(OUTPUT_DIRECTORY).mkdir(parents=True, exist_ok=True)
    
    # Полный путь к Excel файлу
    excel_path = Path(OUTPUT_DIRECTORY) / EXCEL_FILENAME
    
    # Список для хранения информации о файлах
    files_data = []
    
    # Фильтруем существующие директории
    valid_directories = []
    for dir_path in SOURCE_DIRECTORIES:
        if os.path.exists(dir_path):
            valid_directories.append(dir_path)
        else:
            print(f"⚠️ Папка не существует: {dir_path}")
    
    if not valid_directories:
        print(f"❌ Ошибка: Ни одна из указанных папок не существует!")
        return
    
    # Счетчики
    total_processed = 0
    total_files_in_dirs = 0
    matching_files = 0
    
    print("=" * 80)
    print("🔍 АНАЛИЗ ФАЙЛОВ В НЕСКОЛЬКИХ ДИРЕКТОРИЯХ")
    print("=" * 80)
    
    # Выводим информацию о настройках
    print(f"📁 Папки для анализа ({len(valid_directories)}):")
    for i, dir_path in enumerate(valid_directories, 1):
        print(f"   {i}. {dir_path}")
    
    print(f"\n⚙️  Настройки поиска:")
    print(f"   Фильтр по ключевым словам: {'ВКЛЮЧЕН' if ENABLE_KEYWORD_FILTER else 'ВЫКЛЮЧЕН'}")
    
    if ENABLE_KEYWORD_FILTER:
        keywords_str = ", ".join(KEYWORDS) if KEYWORDS else "не заданы"
        print(f"   Ключевые слова: {keywords_str}")
        print(f"   Чувствительность к регистру: {'Да' if CASE_SENSITIVE_SEARCH else 'Нет'}")
        print(f"   Искать только в именах файлов: {'Да' if SEARCH_IN_FILENAME_ONLY else 'Нет'}")
        print(f"   Сохранять все файлы: {'Да' if SAVE_NON_MATCHING_FILES else 'Нет'}")
    
    print(f"🔗 Гиперссылки: {'ВКЛЮЧЕНЫ' if CREATE_HYPERLINKS else 'ВЫКЛЮЧЕНЫ'}")
    print(f"📊 Отчет будет сохранен: {excel_path}")
    print("-" * 80)
    
    # Обрабатываем каждую директорию
    for source_dir in valid_directories:
        print(f"\n📂 Анализируем папку: {source_dir}")
        
        dir_file_count = 0
        dir_matching_count = 0
        
        # Рекурсивно обходим все файлы в директории
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                total_files_in_dirs += 1
                file_path = Path(root) / file
                
                try:
                    # Получаем информацию о файле
                    filename = file_path.name
                    
                    # Тип файла
                    file_extension = file_path.suffix.lower()
                    if file_extension:
                        file_type = file_extension.lstrip('.').upper()
                    else:
                        file_type = "БЕЗ РАСШИРЕНИЯ"
                    
                    # Дата изменения
                    try:
                        mod_time = os.path.getmtime(file_path)
                        mod_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        mod_date = 'НЕДОСТУПНО'
                    
                    # Полный путь
                    full_path = str(file_path.resolve())
                    
                    # Проверяем ключевые слова
                    matched_keywords = []
                    if ENABLE_KEYWORD_FILTER:
                        matched_keywords = check_keywords_match(
                            file_path, filename, KEYWORDS, 
                            CASE_SENSITIVE_SEARCH, SEARCH_IN_FILENAME_ONLY
                        )
                    
                    # Определяем, нужно ли сохранять этот файл
                    should_save = True
                    if ENABLE_KEYWORD_FILTER and not SAVE_NON_MATCHING_FILES:
                        should_save = bool(matched_keywords)
                    
                    if should_save:
                        # Добавляем информацию в список
                        files_data.append([
                            filename,
                            file_type,
                            mod_date,
                            full_path,
                            source_dir,  # Источник (какая папка)
                            matched_keywords
                        ])
                        
                        if matched_keywords:
                            matching_files += 1
                            dir_matching_count += 1
                    
                    total_processed += 1
                    dir_file_count += 1
                    
                    if SHOW_DETAILS and total_processed % 500 == 0:
                        print(f"   📊 Обработано файлов: {total_processed}...")
                        
                except Exception as e:
                    print(f"   ⚠️ Ошибка при обработке файла {file_path}: {e}")
                    continue
        
        print(f"   ✅ Обработано файлов в этой папке: {dir_file_count}")
        if ENABLE_KEYWORD_FILTER:
            print(f"   🔍 Соответствует фильтрам: {dir_matching_count}")
    
    # Создаем Excel отчет
    if files_data:
        print("\n" + "-" * 80)
        print("📈 СОЗДАНИЕ ОТЧЕТА...")
        
        success = create_excel_report(files_data, excel_path, total_processed, matching_files)
        
        print("-" * 80)
        print("🎯 ИТОГОВЫЕ РЕЗУЛЬТАТЫ:")
        print(f"   📄 Всего файлов в папках: {total_files_in_dirs}")
        print(f"   ✅ Успешно обработано: {total_processed}")
        
        if ENABLE_KEYWORD_FILTER:
            print(f"   🔍 Соответствует фильтрам: {matching_files} ({matching_files/total_processed*100:.1f}%)")
            if SAVE_NON_MATCHING_FILES:
                print(f"   📋 Все файлы сохранены в отчет")
            else:
                print(f"   📋 Только соответствующие фильтрам сохранены в отчет")
        
        print(f"   💾 Записей в Excel: {len(files_data)}")
        
        if success:
            print(f"\n   ✅ Excel отчет успешно создан: {excel_path}")
            
            if CREATE_HYPERLINKS:
                print(f"   🔗 Гиперссылки добавлены к именам файлов и путям")
                print(f"   💡 В Excel: щелкните по имени файла или пути для открытия")
            
            # Открываем Excel файл если нужно
            if OPEN_EXCEL_AFTER_CREATION:
                print(f"   📂 Открываю Excel файл...")
                open_excel_file(excel_path)
                
        else:
            print("   ❌ Не удалось создать Excel отчет")
    else:
        print("\nℹ️  Не найдено файлов, соответствующих критериям.")
    
    print("=" * 80)

if __name__ == "__main__":
    analyze_directory_files()
