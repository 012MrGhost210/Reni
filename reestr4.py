import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
import re

# ==================== НАСТРОЙКИ ====================
# Быстро меняйте параметры поиска здесь:

# СПИСОК папок для анализа (можно указать несколько)
SOURCE_DIRECTORIES = [
    r'M:\Финансовый департамент\Treasury',  # ПЕРВАЯ папка
    r'M:\Финансовый департамент',           # ВТОРАЯ папка
    # r'D:\Документы',                      # ТРЕТЬЯ папка (раскомментировать если нужно)
    # Добавьте другие папки по мере необходимости
]

# Папка для сохранения Excel файла
OUTPUT_DIRECTORY = r'\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\Test'

# Название Excel файла с отчетом
EXCEL_FILENAME = "анализ_файлов.xlsx"

# ==================== НАСТРОЙКИ КЛЮЧЕВЫХ СЛОВ ====================

# Включить/выключить фильтрацию по ключевым словам
ENABLE_KEYWORD_FILTER = True  # True - фильтровать по ключевым словам, False - все файлы

# Максимальное количество ключевых слов для отдельных листов
MAX_KEYWORDS_FOR_SHEETS = 10  # Не более 10 ключевых слов получат отдельные листы

# РЕЖИМ ПОИСКА:
# "exact" - точное совпадение слова (целиком, не содержит внутри других слов)
# "contains" - содержит подстроку (старый режим)
SEARCH_MODE = "exact"  # "exact" или "contains"

# ОТКУДА БРАТЬ КЛЮЧЕВЫЕ СЛОВА:
# Вариант 1: Из Excel файла (раскомментируйте и укажите путь)
KEYWORDS_EXCEL_FILE = r'\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\ключевые_слова.xlsx'  # ЗАМЕНИТЕ НА СВОЙ ПУТЬ
KEYWORDS_EXCEL_SHEET = 'Лист1'  # Название листа (по умолчанию 'Лист1')
KEYWORDS_EXCEL_COLUMN = 'A'     # Колонка с ключевыми словами (A, B, C и т.д.)

# Вариант 2: Из списка в коде (раскомментируйте если не используете Excel)
# KEYWORDS = [
#     "Чешенко",
#     "отчет",
#     "финанс",
#     # "2024",  # Добавьте свои ключевые слова
#     # "смета",
# ]

# ==================== НАСТРОЙКИ ПОИСКА ====================

# Чувствительность к регистру при поиске
CASE_SENSITIVE_SEARCH = False  # True - учитывает регистр, False - не учитывает

# Где искать ключевые слова
SEARCH_IN_FILENAME_ONLY = False  # True - только в имени файла, False - в полном пути

# ==================== ДОПОЛНИТЕЛЬНЫЕ НАСТРОЙКИ ====================

# Показывать подробный процесс работы
SHOW_DETAILS = True

# Создать папку для отчета, если её нет
CREATE_OUTPUT_DIR = True

# Делать ли гиперссылки на файлы
CREATE_HYPERLINKS = True

# Открывать ли Excel файл после создания
OPEN_EXCEL_AFTER_CREATION = True

# ==================== КОНЕЦ НАСТРОЕК ====================

def load_keywords_from_excel():
    """
    Загружает ключевые слова из Excel файла
    Возвращает список ключевых слов или пустой список в случае ошибки
    """
    try:
        if not KEYWORDS_EXCEL_FILE or not os.path.exists(KEYWORDS_EXCEL_FILE):
            print(f"⚠️ Excel файл с ключевыми словами не найден: {KEYWORDS_EXCEL_FILE}")
            return []
        
        print(f"📖 Загружаем ключевые слова из: {KEYWORDS_EXCEL_FILE}")
        
        # Определяем букву колонки для pandas
        col_letter = KEYWORDS_EXCEL_COLUMN.upper()
        col_index = ord(col_letter) - 65  # A=0, B=1, C=2 и т.д.
        
        # Читаем Excel файл
        df = pd.read_excel(KEYWORDS_EXCEL_FILE, sheet_name=KEYWORDS_EXCEL_SHEET, header=None)
        
        # Получаем значения из указанной колонки
        keywords = []
        for idx, row in df.iterrows():
            if col_index < len(row):
                value = row[col_index]
                # Проверяем, что значение не пустое и это строка
                if pd.notna(value) and isinstance(value, (str, int, float)):
                    # Приводим к строке и удаляем лишние пробелы
                    keyword = str(value).strip()
                    if keyword:  # Не пустая строка
                        keywords.append(keyword)
        
        # Убираем дубликаты и сортируем
        keywords = sorted(list(set(keywords)))
        
        print(f"✅ Загружено ключевых слов: {len(keywords)}")
        if SHOW_DETAILS and keywords:
            print(f"   Примеры: {', '.join(keywords[:10])}{'...' if len(keywords) > 10 else ''}")
        
        return keywords
        
    except Exception as e:
        print(f"❌ Ошибка при загрузке ключевых слов из Excel: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_keywords():
    """
    Возвращает список ключевых слов в зависимости от настроек
    """
    # Пробуем загрузить из Excel
    excel_keywords = load_keywords_from_excel()
    if excel_keywords:
        return excel_keywords
    
    # Если не получилось загрузить из Excel, используем список из кода
    try:
        if KEYWORDS and isinstance(KEYWORDS, list):
            print("📝 Используем ключевые слова из кода")
            return KEYWORDS
    except NameError:
        # Переменная KEYWORDS не определена
        pass
    
    print("⚠️ Нет ключевых слов для поиска!")
    return []

def check_keyword_match_exact(text, keyword, case_sensitive):
    """
    Проверяет точное совпадение слова (не является частью другого слова)
    """
    if not case_sensitive:
        text = text.lower()
        keyword = keyword.lower()
    
    # Создаем паттерн для поиска целого слова
    # \b - граница слова в регулярных выражениях
    pattern = r'\b' + re.escape(keyword) + r'\b'
    return bool(re.search(pattern, text))

def check_keyword_match_contains(text, keyword, case_sensitive):
    """
    Проверяет, содержит ли текст подстроку (старый режим)
    """
    if not case_sensitive:
        text = text.lower()
        keyword = keyword.lower()
    
    return keyword in text

def check_keywords_match(file_path, filename, keywords, case_sensitive, 
                        search_in_filename_only, search_mode):
    """
    Проверяет, содержит ли файл/путь указанные ключевые слова
    Возвращает список найденных ключевых слов
    """
    matched_keywords = []
    
    if not keywords or not ENABLE_KEYWORD_FILTER:
        return []
    
    # Выбираем текст для поиска
    search_text = filename if search_in_filename_only else str(file_path)
    
    # Выбираем функцию поиска в зависимости от режима
    if search_mode == "exact":
        match_func = check_keyword_match_exact
    else:  # contains
        match_func = check_keyword_match_contains
    
    for keyword in keywords:
        if not keyword or not keyword.strip():
            continue
        
        if match_func(search_text, keyword, case_sensitive):
            matched_keywords.append(keyword)
    
    return matched_keywords

def format_worksheet(worksheet, title, has_data=True):
    """
    Форматирует отдельный лист Excel
    """
    # Устанавливаем ширину столбцов
    column_widths = {
        'A': 40,   # Имя файла
        'B': 20,   # Тип файла
        'C': 25,   # Дата изменения
        'D': 100,  # Полный путь
        'E': 30,   # Источник (папка поиска)
    }
    
    for col, width in column_widths.items():
        col_letter = col
        worksheet.column_dimensions[col_letter].width = width
    
    # Форматируем заголовки
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    headers = ["Имя файла", "Тип файла", "Дата изменения", "Полный путь", "Источник (папка)"]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if has_data:
        # Делаем автофильтр
        worksheet.auto_filter.ref = f"A1:E{worksheet.max_row}"
        
        # Настраиваем стиль для гиперссылок
        hyperlink_font = Font(color="0563C1", underline="single")
        
        # Применяем стиль гиперссылок
        for row in range(2, worksheet.max_row + 1):
            path_cell = worksheet.cell(row=row, column=4)
            
            if CREATE_HYPERLINKS and path_cell.hyperlink:
                path_cell.font = hyperlink_font
                
                name_cell = worksheet.cell(row=row, column=1)
                name_cell.hyperlink = path_cell.hyperlink
                name_cell.font = hyperlink_font
        
        # Условное форматирование для дат
        date_column_letter = 'C'
        date_range = f"{date_column_letter}2:{date_column_letter}{worksheet.max_row}"
        
        today_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        today_font = Font(color="006100")
        
        today_rule = FormulaRule(
            formula=[f'AND(${date_column_letter}2>=TODAY(), ${date_column_letter}2<TODAY()+1)'],
            fill=today_fill,
            font=today_font
        )
        worksheet.conditional_formatting.add(date_range, today_rule)
        
        old_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        old_font = Font(color="9C0006")
        
        old_rule = FormulaRule(
            formula=[f'${date_column_letter}2<TODAY()-30'],
            fill=old_fill,
            font=old_font
        )
        worksheet.conditional_formatting.add(date_range, old_rule)
    
    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"

def create_file_hyperlink(file_path):
    """
    Создает корректную гиперссылку для файла
    """
    try:
        if not os.path.exists(file_path):
            return None
        
        abs_path = os.path.abspath(file_path)
        hyperlink_path = abs_path.replace('/', '\\')
        
        return hyperlink_path
    except:
        return None

def should_save_file(matched_keywords, enable_filter):
    """
    Определяет, нужно ли сохранять файл в отчете
    """
    if not enable_filter:
        return True
    else:
        return bool(matched_keywords)

def create_excel_report(files_data, all_files_data, keywords_list, output_path, 
                       total_processed, matching_files, search_mode):
    """
    Создает Excel файл с отдельными листами для каждого ключевого слова
    """
    try:
        workbook = Workbook()
        
        # Удаляем стандартный лист
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # 1. Лист "Все файлы" (всегда)
        ws_all = workbook.create_sheet("Все файлы")
        all_files_rows = []
        
        for file_info in all_files_data:
            filename, file_type, mod_date, full_path, source_dir, matched = file_info
            all_files_rows.append([filename, file_type, mod_date, full_path, source_dir])
        
        # Заполняем лист "Все файлы"
        for row_idx, row_data in enumerate(all_files_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_all.cell(row=row_idx, column=col_idx, value=value)
                
                # Добавляем гиперссылку для полного пути
                if col_idx == 4 and CREATE_HYPERLINKS:
                    hyperlink = create_file_hyperlink(value)
                    if hyperlink:
                        ws_all.cell(row=row_idx, column=col_idx).hyperlink = hyperlink
        
        format_worksheet(ws_all, "Все файлы", bool(all_files_rows))
        
        # Добавляем информацию о количестве на лист "Все файлы"
        info_row = len(all_files_rows) + 3
        ws_all.cell(row=info_row, column=1, value=f"Всего файлов: {len(all_files_rows)}")
        ws_all.cell(row=info_row, column=1).font = Font(bold=True, size=12)
        
        # 2. Если фильтр включен и есть ключевые слова - создаем отдельные листы
        if ENABLE_KEYWORD_FILTER and keywords_list:
            # Берем не более MAX_KEYWORDS_FOR_SHEETS ключевых слов
            keywords_for_sheets = keywords_list[:MAX_KEYWORDS_FOR_SHEETS]
            
            print(f"\n📑 Создаем отдельные листы для {len(keywords_for_sheets)} ключевых слов:")
            
            # Словарь для хранения файлов по ключевым словам
            keyword_files = {kw: [] for kw in keywords_for_sheets}
            
            # Распределяем файлы по ключевым словам
            for file_info in files_data:
                filename, file_type, mod_date, full_path, source_dir, matched_keywords = file_info
                
                for keyword in matched_keywords:
                    if keyword in keyword_files:
                        keyword_files[keyword].append([
                            filename, file_type, mod_date, full_path, source_dir
                        ])
            
            # Создаем листы для каждого ключевого слова
            for keyword in keywords_for_sheets:
                # Очищаем название листа от недопустимых символов
                sheet_name = re.sub(r'[\[\]\*\?\/\\]', '_', keyword)[:31]  # Макс 31 символ для Excel
                
                ws = workbook.create_sheet(sheet_name)
                files_for_keyword = keyword_files[keyword]
                
                # Заполняем лист
                for row_idx, row_data in enumerate(files_for_keyword, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # Добавляем гиперссылку для полного пути
                        if col_idx == 4 and CREATE_HYPERLINKS:
                            hyperlink = create_file_hyperlink(value)
                            if hyperlink:
                                ws.cell(row=row_idx, column=col_idx).hyperlink = hyperlink
                
                format_worksheet(ws, sheet_name, bool(files_for_keyword))
                
                # Добавляем информацию о количестве
                info_row = len(files_for_keyword) + 3
                ws.cell(row=info_row, column=1, value=f"Найдено файлов: {len(files_for_keyword)}")
                ws.cell(row=info_row, column=1).font = Font(bold=True, size=12)
                
                print(f"   📄 {keyword}: {len(files_for_keyword)} файлов")
            
            # Если ключевых слов больше MAX_KEYWORDS_FOR_SHEETS, создаем дополнительный лист
            if len(keywords_list) > MAX_KEYWORDS_FOR_SHEETS:
                ws_extra = workbook.create_sheet("Остальные ключевые слова")
                extra_keywords = keywords_list[MAX_KEYWORDS_FOR_SHEETS:]
                
                row_idx = 2
                for keyword in extra_keywords:
                    ws_extra.cell(row=row_idx, column=1, value=keyword)
                    row_idx += 1
                
                ws_extra.column_dimensions['A'].width = 40
                ws_extra.cell(row=1, column=1, value="Ключевые слова").font = Font(bold=True)
                ws_extra.freeze_panes = "A2"
                
                print(f"   📄 Остальные ключевые слова: {len(extra_keywords)} шт.")
        
        # 3. Лист со статистикой
        ws_stats = workbook.create_sheet("Статистика")
        
        stats_data = [
            ["Параметр", "Значение"],
            ["Дата создания отчета", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Всего обработано файлов", total_processed],
            ["Всего файлов в отчете", len(all_files_data)],
            ["Режим фильтрации", "ВКЛЮЧЕН" if ENABLE_KEYWORD_FILTER else "ВЫКЛЮЧЕН"],
        ]
        
        if ENABLE_KEYWORD_FILTER:
            stats_data.extend([
                ["Режим поиска", "точное совпадение" if search_mode == "exact" else "поиск подстроки"],
                ["Найдено файлов с ключевыми словами", matching_files],
                ["Всего ключевых слов", len(keywords_list)],
                ["Ключевых слов с отдельными листами", min(len(keywords_list), MAX_KEYWORDS_FOR_SHEETS)],
            ])
            
            # Добавляем список ключевых слов
            stats_data.append(["", ""])
            stats_data.append(["СПИСОК КЛЮЧЕВЫХ СЛОВ:", ""])
            for i, kw in enumerate(keywords_list, 1):
                stats_data.append([f"{i}. {kw}", ""])
        
        stats_data.extend([
            ["", ""],
            ["Настройки поиска:", ""],
            ["Чувствительность к регистру", "Да" if CASE_SENSITIVE_SEARCH else "Нет"],
            ["Поиск только в именах файлов", "Да" if SEARCH_IN_FILENAME_ONLY else "Нет"],
            ["Гиперссылки", "Да" if CREATE_HYPERLINKS else "Нет"],
        ])
        
        # Заполняем статистику
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Заголовок
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 50
        
        # Сохраняем файл
        workbook.save(output_path)
        return True
        
    except Exception as e:
        print(f"❌ Ошибка при создании Excel отчета: {e}")
        import traceback
        traceback.print_exc()
        return False

def open_excel_file(file_path):
    try:
        os.startfile(file_path)
        return True
    except Exception as e:
        print(f"⚠️ Не удалось открыть Excel файл: {e}")
        return False

def analyze_directory_files():
    """
    Анализирует файлы в нескольких директориях с фильтрацией по ключевым словам
    """
    if CREATE_OUTPUT_DIR:
        Path(OUTPUT_DIRECTORY).mkdir(parents=True, exist_ok=True)
    
    excel_path = Path(OUTPUT_DIRECTORY) / EXCEL_FILENAME
    
    # Загружаем ключевые слова
    keywords_list = []
    if ENABLE_KEYWORD_FILTER:
        keywords_list = get_keywords()
    
    # Фильтруем существующие директории
    valid_directories = []
    for dir_path in SOURCE_DIRECTORIES:
        if os.path.exists(dir_path):
            valid_directories.append(dir_path)
        else:
            print(f"⚠️ Папка не существует: {dir_path}")
    
    if not valid_directories:
        print(f"❌ Ошибка: Ни одна из указанных папок не существует!")
        return
    
    # Списки для хранения данных
    all_files_data = []      # Все найденные файлы
    filtered_files_data = [] # Только файлы с ключевыми словами (если фильтр включен)
    
    # Счетчики
    total_processed = 0
    total_files_in_dirs = 0
    matching_files = 0
    
    print("=" * 80)
    print("🔍 АНАЛИЗ ФАЙЛОВ С ОТДЕЛЬНЫМИ ЛИСТАМИ ПО КЛЮЧЕВЫМ СЛОВАМ")
    print("=" * 80)
    
    print(f"📁 Папки для анализа ({len(valid_directories)}):")
    for i, dir_path in enumerate(valid_directories, 1):
        print(f"   {i}. {dir_path}")
    
    print(f"\n⚙️  Настройки поиска:")
    print(f"   Фильтр по ключевым словам: {'ВКЛЮЧЕН' if ENABLE_KEYWORD_FILTER else 'ВЫКЛЮЧЕН'}")
    
    if ENABLE_KEYWORD_FILTER:
        mode_text = "ТОЧНОЕ СОВПАДЕНИЕ СЛОВ" if SEARCH_MODE == "exact" else "ПОИСК ПОДСТРОКИ"
        print(f"   Режим поиска: {mode_text}")
        print(f"   Отдельные листы для первых {MAX_KEYWORDS_FOR_SHEETS} ключевых слов")
        
        if keywords_list:
            print(f"   Ключевые слова: {len(keywords_list)} шт.")
            print(f"   Первые 10: {', '.join(keywords_list[:10])}{'...' if len(keywords_list) > 10 else ''}")
        else:
            print(f"   ⚠️ Ключевые слова не загружены!")
    
    print(f"🔗 Гиперссылки: {'ВКЛЮЧЕНЫ' if CREATE_HYPERLINKS else 'ВЫКЛЮЧЕНЫ'}")
    print(f"📊 Отчет будет сохранен: {excel_path}")
    print("-" * 80)
    
    # Обрабатываем каждую директорию
    for source_dir in valid_directories:
        print(f"\n📂 Анализируем папку: {source_dir}")
        
        dir_file_count = 0
        dir_matching_count = 0
        
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                total_files_in_dirs += 1
                file_path = Path(root) / file
                
                try:
                    filename = file_path.name
                    
                    file_extension = file_path.suffix.lower()
                    if file_extension:
                        file_type = file_extension.lstrip('.').upper()
                    else:
                        file_type = "БЕЗ РАСШИРЕНИЯ"
                    
                    try:
                        mod_time = os.path.getmtime(file_path)
                        mod_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        mod_date = 'НЕДОСТУПНО'
                    
                    full_path = str(file_path.resolve())
                    
                    # Проверяем ключевые слова
                    matched_keywords = []
                    if ENABLE_KEYWORD_FILTER and keywords_list:
                        matched_keywords = check_keywords_match(
                            file_path, filename, keywords_list, 
                            CASE_SENSITIVE_SEARCH, SEARCH_IN_FILENAME_ONLY,
                            SEARCH_MODE
                        )
                    
                    # Сохраняем в общий список (ВСЕ файлы)
                    all_files_data.append([
                        filename,
                        file_type,
                        mod_date,
                        full_path,
                        source_dir,
                        matched_keywords
                    ])
                    
                    # Если есть совпадения, сохраняем в отфильтрованный список
                    if matched_keywords:
                        filtered_files_data.append([
                            filename,
                            file_type,
                            mod_date,
                            full_path,
                            source_dir,
                            matched_keywords
                        ])
                        matching_files += 1
                        dir_matching_count += 1
                    
                    total_processed += 1
                    dir_file_count += 1
                    
                    if SHOW_DETAILS and total_processed % 500 == 0:
                        print(f"   📊 Обработано файлов: {total_processed}...")
                        
                except Exception as e:
                    print(f"   ⚠️ Ошибка при обработке файла {file_path}: {e}")
                    continue
        
        print(f"   ✅ Обработано файлов в этой папке: {dir_file_count}")
        if ENABLE_KEYWORD_FILTER and keywords_list:
            print(f"   🔍 Найдено с ключевыми словами: {dir_matching_count}")
    
    # Создаем Excel отчет
    if all_files_data:
        print("\n" + "-" * 80)
        print("📈 СОЗДАНИЕ ОТЧЕТА С ОТДЕЛЬНЫМИ ЛИСТАМИ...")
        
        # Выбираем какие данные использовать для фильтрации
        files_for_filter = filtered_files_data if ENABLE_KEYWORD_FILTER else []
        
        success = create_excel_report(
            files_for_filter,      # Только файлы с ключевыми словами (для отдельных листов)
            all_files_data,        # Все файлы (для листа "Все файлы")
            keywords_list,         # Список ключевых слов
            excel_path, 
            total_processed, 
            matching_files,
            SEARCH_MODE
        )
        
        print("-" * 80)
        print("🎯 ИТОГОВЫЕ РЕЗУЛЬТАТЫ:")
        print(f"   📄 Всего файлов в папках: {total_files_in_dirs}")
        print(f"   ✅ Успешно обработано: {total_processed}")
        print(f"   📋 Всего записей в отчете: {len(all_files_data)}")
        
        if ENABLE_KEYWORD_FILTER:
            print(f"   🔍 Найдено с ключевыми словами: {matching_files}")
            if keywords_list:
                sheets_count = min(len(keywords_list), MAX_KEYWORDS_FOR_SHEETS)
                print(f"   📑 Создано листов по ключевым словам: {sheets_count}")
                if len(keywords_list) > MAX_KEYWORDS_FOR_SHEETS:
                    print(f"   📑 Остальные ключевые слова: {len(keywords_list) - MAX_KEYWORDS_FOR_SHEETS} шт. на отдельном листе")
        
        if success:
            print(f"\n   ✅ Excel отчет успешно создан: {excel_path}")
            
            if CREATE_HYPERLINKS:
                print(f"   🔗 Гиперссылки добавлены к именам файлов и путям")
            
            if OPEN_EXCEL_AFTER_CREATION:
                print(f"   📂 Открываю Excel файл...")
                open_excel_file(excel_path)
        else:
            print("   ❌ Не удалось создать Excel отчет")
    else:
        print("\nℹ️  Не найдено файлов.")
    
    print("=" * 80)

if __name__ == "__main__":
    analyze_directory_files()
