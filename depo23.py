import pandas as pd
import tkinter as tk
from tkinter import messagebox, ttk
import os
from datetime import datetime

# ФИКСИРОВАННЫЕ ПУТИ К ФАЙЛАМ
FILE1_PATH = r"M:\Финансовый департамент\Treasury\Базы данных(автоматизация)\DI_DATABASE\Депозит.xlsx"
REGISTRY_FILE_PATH = r"\\fs-01.renlife.com\alldocs\Инвестиционный департамент\7.0 Treasury\BigData\Deposit.xlsx"

# Список контрагентов для отслеживания (можно добавить других)
TRACKED_CONTRACTORS = ["ПОЧТА РОССИИ", "Почта России"]

def parse_number(text):
    """Преобразует строку с числами в формате '12 122 121,31' в float"""
    if isinstance(text, (int, float)):
        return float(text)
    # Убираем пробелы и заменяем запятую на точку
    cleaned = str(text).strip().replace(' ', '').replace(',', '.')
    # Если строка пустая или не число, возвращаем 0
    if not cleaned:
        return 0.0
    return float(cleaned)

def format_number(number):
    """Форматирует число в строку с пробелами тысяч и запятой для десятичных"""
    return f"{number:,.2f}".replace(',', ' ').replace('.', ',')

def calculate_days_until(target_date_str):
    """Вычисляет количество дней от сегодняшней даты до целевой даты"""
    today = datetime.now().date()
    
    # Парсим дату из разных возможных форматов
    for fmt in ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%y']:
        try:
            target_date = datetime.strptime(target_date_str, fmt).date()
            days = (target_date - today).days
            return max(days, 1)  # Минимум 1 день
        except ValueError:
            continue
    
    # Если не удалось распарсить, возвращаем исходную строку
    return target_date_str

def get_tracked_payments():
    """Читает файл и находит платежи от отслеживаемых контрагентов"""
    try:
        if not os.path.exists(FILE1_PATH):
            return []
        
        df = pd.read_excel(FILE1_PATH)
        if df.empty:
            return []
        
        tracked_payments = []
        
        # Столбец I - контрагенты, столбец H - суммы
        # Предполагаем, что столбцы: H (индекс 7) - суммы, I (индекс 8) - контрагенты
        for idx, row in df.iterrows():
            if len(row) > 8:  # Проверяем, что есть столбец I
                contractor = str(row.iloc[8]) if pd.notna(row.iloc[8]) else ""
                
                # Проверяем, содержит ли наименование контрагента отслеживаемое слово
                for tracked in TRACKED_CONTRACTORS:
                    if tracked in contractor.upper():
                        # Нашли отслеживаемого контрагента
                        amount = 0
                        if len(row) > 7 and pd.notna(row.iloc[7]):
                            amount = parse_number(row.iloc[7])
                        
                        tracked_payments.append({
                            'contractor': contractor,
                            'amount': amount,
                            'row': idx + 2  # +2 для учета заголовков и индексации от 1
                        })
                        break
        
        return tracked_payments
        
    except Exception as e:
        print(f"Ошибка при поиске платежей: {e}")
        return []

def save_to_registry(deposits_data, total_amount):
    """Сохраняет данные о депозитах в реестр"""
    try:
        # Открываем файл для записи с помощью openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        from datetime import datetime, date
        
        if os.path.exists(REGISTRY_FILE_PATH):
            wb = load_workbook(REGISTRY_FILE_PATH)
        else:
            messagebox.showerror("Ошибка", f"Файл реестра не найден по пути:\n{REGISTRY_FILE_PATH}")
            return False
        
        # Выбираем лист "2024-2025"
        if '2024-2025' in wb.sheetnames:
            ws = wb['2024-2025']
        else:
            messagebox.showerror("Ошибка", "Лист '2024-2025' не найден в файле реестра")
            return False
        
        # Находим первую пустую строку в столбце D (сумма)
        current_row = 4  # Начинаем с 4 строки (после заголовков)
        while ws.cell(row=current_row, column=4).value is not None:  # Столбец D
            current_row += 1
        
        # Запоминаем первую строку для применения формата
        first_row = current_row
        
        today_date = date.today()  # Используем date вместо datetime
        
        # Записываем данные для каждого депозита
        for i, deposit in enumerate(deposits_data):
            # Столбец D (4) - Сумма
            ws.cell(row=current_row, column=4, value=float(deposit['amount']))
            
            # Столбец E (5) - Ставка
            rate_str = deposit['rate'].replace('%', '').replace(',', '.')
            rate_value = float(rate_str) / 100
            ws.cell(row=current_row, column=5, value=rate_value)
            
            # Столбец H (8) - Размещение (сегодня) - ТОЛЬКО ДАТА
            ws.cell(row=current_row, column=8, value=today_date)
            
            # Столбец I (9) - Окончание - ТОЛЬКО ДАТА
            # Преобразуем строку с датой
            date_str = deposit['date'].strip()
            
            try:
                # Разбиваем строку на день, месяц, год
                day, month, year = map(int, date_str.split('.'))
                date_obj = date(year, month, day)  # Используем date вместо datetime
                
                # Записываем как дату
                ws.cell(row=current_row, column=9, value=date_obj)
                # Устанавливаем формат ячейки как дата
                ws.cell(row=current_row, column=9).number_format = 'DD.MM.YYYY'
                
            except Exception as e:
                # Если не удалось распарсить, записываем как строку
                print(f"Не удалось распарсить дату {date_str}: {e}")
                ws.cell(row=current_row, column=9, value=date_str)
            
            current_row += 1
        
        # Применяем формат даты ко всем записанным ячейкам
        for row in range(first_row, current_row):
            # Формат для столбца H (размещение)
            cell_h = ws.cell(row=row, column=8)
            cell_h.number_format = 'DD.MM.YYYY'
            
            # Формат для столбца I (окончание)
            cell_i = ws.cell(row=row, column=9)
            if isinstance(cell_i.value, (datetime, date)):
                cell_i.number_format = 'DD.MM.YYYY'
        
        # Сохраняем файл
        wb.save(REGISTRY_FILE_PATH)
        
        # Показываем сообщение об успехе
        messagebox.showinfo("Успех", 
                          f"Данные о депозитах успешно сохранены в реестр!\n"
                          f"Записано депозитов: {len(deposits_data)}")
        return True
        
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при сохранении в реестр: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def update_available_balance():
    """Обновляет отображение доступной суммы на основе текущего баланса и платежей"""
    if not hasattr(root, 'initial_balance'):
        if 'available_balance_label' in globals() and available_balance_label:
            available_balance_label.config(text="Доступно для размещения: сначала загрузите баланс")
        return
    
    try:
        payment_text = entry_payment.get().strip()
        payment_amount = parse_number(payment_text) if payment_text else 0
    except:
        payment_amount = 0
    
    available = root.initial_balance - payment_amount - 10000
    available = max(0, available)
    
    if 'available_balance_label' in globals() and available_balance_label:
        available_balance_label.config(text=f"Доступно для размещения: {format_number(available)}")
    
    # Обновляем остаток после распределения
    update_remaining_balance()
    
    return available

def update_remaining_balance():
    """Обновляет отображение остатка после введенных сумм депозитов"""
    if not hasattr(root, 'initial_balance'):
        return
    
    try:
        payment_text = entry_payment.get().strip()
        payment_amount = parse_number(payment_text) if payment_text else 0
    except:
        payment_amount = 0
    
    available = root.initial_balance - payment_amount - 10000
    available = max(0, available)
    
    # Суммируем введенные суммы депозитов
    total_deposits = 0
    if deposit_entries:
        for deposit in deposit_entries:
            try:
                amount_text = deposit['amount'].get().strip()
                if amount_text:
                    total_deposits += parse_number(amount_text)
            except:
                pass
    
    remaining = available - total_deposits
    remaining = max(0, remaining)
    
    if 'remaining_label' in globals() and remaining_label:
        if remaining == 0 and total_deposits > 0:
            remaining_label.config(text=f"Остаток: {format_number(remaining)} (✓ всё распределено)", fg="green")
        elif remaining > 0:
            remaining_label.config(text=f"Остаток для распределения: {format_number(remaining)}", fg="blue")
        elif total_deposits > available:
            over = total_deposits - available
            remaining_label.config(text=f"Перебор: {format_number(over)}", fg="red")

def auto_fill_deposits(mode='equal'):
    """Автоматически заполняет суммы депозитов"""
    if not hasattr(root, 'initial_balance'):
        messagebox.showerror("Ошибка", "Сначала загрузите баланс (нажмите 'Сформировать сообщение')")
        return
    
    try:
        payment_text = entry_payment.get().strip()
        if not payment_text:
            messagebox.showerror("Ошибка", "Сначала введите сумму платежей сегодня")
            return
        payment_amount = parse_number(payment_text)
    except:
        messagebox.showerror("Ошибка", "Введите корректную сумму платежей")
        return
    
    available = root.initial_balance - payment_amount - 10000
    available = max(0, available)
    
    if available <= 0:
        messagebox.showerror("Ошибка", f"Нет доступных средств!\nБаланс: {format_number(root.initial_balance)}\nПлатежи: {format_number(payment_amount)}")
        return
    
    count = len(deposit_entries)
    
    if mode == 'equal':
        # Распределяем поровну (с округлением до тысяч вниз)
        amount_per_deposit = int(available / count / 1000) * 1000
        remainder = available - (amount_per_deposit * count)
        
        for i, deposit in enumerate(deposit_entries):
            if i == 0 and remainder > 0:
                # Первому депозиту добавляем остаток
                amount = amount_per_deposit + remainder
            else:
                amount = amount_per_deposit
            
            deposit['amount'].delete(0, tk.END)
            deposit['amount'].insert(0, str(amount).replace('.', ','))
        
        messagebox.showinfo("Готово", 
                           f"Сумма {format_number(available)} распределена поровну между {count} депозитами\n"
                           f"По {format_number(amount_per_deposit)} на каждый\n"
                           f"Первому депозиту добавлен остаток: {format_number(remainder)}")
        
    elif mode == 'main':
        # Основной депозит получает все, остальные по 1000
        main_amount = available - (1000 * (count - 1))
        main_amount = max(0, int(main_amount / 1000) * 1000)  # Округляем до тысяч вниз
        
        if main_amount < 0:
            messagebox.showerror("Ошибка", "Сумма слишком мала для распределения")
            return
        
        for i, deposit in enumerate(deposit_entries):
            if i == 0:
                amount = main_amount
            else:
                amount = 1000
            
            deposit['amount'].delete(0, tk.END)
            deposit['amount'].insert(0, str(amount).replace('.', ','))
        
        messagebox.showinfo("Готово", 
                           f"Сумма {format_number(available)} распределена:\n"
                           f"Депозит 1 (основной): {format_number(main_amount)}\n"
                           f"Остальные {count-1} депозитов: по 1 000")
    
    # Обновляем остаток
    update_remaining_balance()

def on_deposit_count_change(*args):
    """Обработчик изменения количества депозитов"""
    count = deposit_count.get()
    
    # Очищаем старые поля
    for widget in deposit_fields_frame.winfo_children():
        widget.destroy()
    
    # Всегда показываем поле для суммы платежей
    tk.Label(deposit_fields_frame, text="Сумма платежей сегодня:", font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 5))
    
    # Создаем фрейм для платежей
    payments_frame = tk.Frame(deposit_fields_frame)
    payments_frame.pack(fill=tk.X, pady=(0, 10))
    
    global entry_payment
    entry_payment = tk.Entry(payments_frame, width=30)
    entry_payment.pack(side=tk.LEFT, padx=(0, 10))
    
    # Кнопка для поиска платежей от контрагентов
    btn_find_payments = tk.Button(payments_frame, text="Добавить платежи от контрагентов", 
                                  command=add_tracked_payments, bg="#FF9800", fg="white")
    btn_find_payments.pack(side=tk.LEFT)
    
    if count == "1":
        # Поля для одного депозита
        tk.Label(deposit_fields_frame, text="Ставка (%):").pack(anchor=tk.W)
        global entry_rate
        entry_rate = tk.Entry(deposit_fields_frame, width=30)
        entry_rate.pack(fill=tk.X, pady=(5, 10))
        
        tk.Label(deposit_fields_frame, text="Срок до (дата):").pack(anchor=tk.W)
        tk.Label(deposit_fields_frame, text="Формат: ДД.ММ.ГГГГ (например: 31.12.2024)", 
                 fg="gray", font=("Arial", 8)).pack(anchor=tk.W)
        global entry_date
        entry_date = tk.Entry(deposit_fields_frame, width=30)
        entry_date.pack(fill=tk.X, pady=(5, 10))
        
        # Скрываем переменные для нескольких депозитов, если они есть
        if 'available_balance_label' in globals() and available_balance_label:
            available_balance_label.pack_forget()
        if 'remaining_label' in globals() and remaining_label:
            remaining_label.pack_forget()
        
    else:
        # Поля для нескольких депозитов
        count_num = int(count)
        
        # Создаем разделитель
        tk.Frame(deposit_fields_frame, height=2, bg="gray").pack(fill=tk.X, pady=10)
        
        # Фрейм для отображения доступного остатка
        balance_frame = tk.Frame(deposit_fields_frame, bg="#E3F2FD", relief=tk.GROOVE, bd=1)
        balance_frame.pack(fill=tk.X, pady=(0, 10))
        
        global available_balance_label, remaining_label
        available_balance_label = tk.Label(balance_frame, text="Доступно для размещения: 0", 
                                           font=("Arial", 11, "bold"), fg="green", bg="#E3F2FD")
        available_balance_label.pack(pady=(5, 2))
        
        remaining_label = tk.Label(balance_frame, text="Остаток после распределения: 0", 
                                   font=("Arial", 10), fg="blue", bg="#E3F2FD")
        remaining_label.pack(pady=(0, 5))
        
        # Кнопки для автоматического заполнения
        auto_frame = tk.Frame(deposit_fields_frame)
        auto_frame.pack(fill=tk.X, pady=(0, 10))
        
        btn_equal = tk.Button(auto_frame, text="📊 Распределить поровну", 
                             command=lambda: auto_fill_deposits('equal'),
                             bg="#4CAF50", fg="white", font=("Arial", 9, "bold"))
        btn_equal.pack(side=tk.LEFT, padx=(0, 10))
        
        btn_main = tk.Button(auto_frame, text="⭐ Основной + остальные по 1000", 
                            command=lambda: auto_fill_deposits('main'),
                            bg="#FF9800", fg="white", font=("Arial", 9, "bold"))
        btn_main.pack(side=tk.LEFT)
        
        tk.Label(deposit_fields_frame, text=f"Данные для {count_num} депозитов:", 
                 font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(10, 10))
        
        # Создаем поля для каждого депозита
        global deposit_entries
        deposit_entries = []
        
        # Фрейм для списка депозитов
        deposits_container = tk.Frame(deposit_fields_frame)
        deposits_container.pack(fill=tk.X)
        
        for i in range(count_num):
            dep_frame = tk.Frame(deposits_container, relief=tk.GROOVE, bd=1, padx=10, pady=8)
            dep_frame.pack(fill=tk.X, pady=5)
            
            tk.Label(dep_frame, text=f"Депозит {i+1}", font=("Arial", 9, "bold")).pack(anchor=tk.W)
            
            # Сумма депозита
            tk.Label(dep_frame, text="Сумма:").pack(anchor=tk.W)
            entry_amount_dep = tk.Entry(dep_frame, width=30)
            entry_amount_dep.pack(fill=tk.X, pady=(2, 5))
            # Привязываем событие изменения для пересчета остатка
            entry_amount_dep.bind('<KeyRelease>', lambda e, idx=i: update_remaining_balance())
            
            # Ставка
            tk.Label(dep_frame, text="Ставка (%):").pack(anchor=tk.W)
            entry_rate_dep = tk.Entry(dep_frame, width=30)
            entry_rate_dep.pack(fill=tk.X, pady=(2, 5))
            
            # Дата
            tk.Label(dep_frame, text="Срок до (дата):").pack(anchor=tk.W)
            tk.Label(dep_frame, text="Формат: ДД.ММ.ГГГГ", fg="gray", font=("Arial", 7)).pack(anchor=tk.W)
            entry_date_dep = tk.Entry(dep_frame, width=30)
            entry_date_dep.pack(fill=tk.X, pady=(2, 5))
            
            deposit_entries.append({
                'amount': entry_amount_dep,
                'rate': entry_rate_dep,
                'date': entry_date_dep,
                'frame': dep_frame
            })
        
        # Добавляем кнопку обновления остатка
        btn_refresh = tk.Button(deposit_fields_frame, text="🔄 Обновить доступную сумму", 
                               command=update_available_balance,
                               bg="#2196F3", fg="white")
        btn_refresh.pack(pady=(10, 0))
        
        # Изначально показываем остаток (если баланс уже загружен)
        if hasattr(root, 'initial_balance'):
            update_available_balance()

def add_tracked_payments():
    """Находит платежи от отслеживаемых контрагентов и ДОБАВЛЯЕТ к уже введенной сумме"""
    tracked_payments = get_tracked_payments()
    
    if not tracked_payments:
        messagebox.showinfo("Информация", "Платежи от отслеживаемых контрагентов не найдены")
        return
    
    # Суммируем все найденные платежи
    total_found = sum(p['amount'] for p in tracked_payments)
    
    # Формируем сообщение о найденных платежах
    payments_info = f"Найдено платежей: {len(tracked_payments)}\n\n"
    for payment in tracked_payments:
        payments_info += f"• {payment['contractor'][:50]}: {format_number(payment['amount'])}\n"
    payments_info += f"\nОбщая сумма найденных платежей: {format_number(total_found)}"
    
    # Получаем текущую сумму из поля
    current_text = entry_payment.get().strip()
    current_amount = 0
    if current_text:
        try:
            current_amount = parse_number(current_text)
        except:
            pass
    
    new_amount = current_amount + total_found
    
    # Спрашиваем пользователя, хочет ли он добавить эту сумму
    result = messagebox.askyesno("Найдены платежи", 
                                 f"{payments_info}\n\n"
                                 f"Текущая сумма: {format_number(current_amount)}\n"
                                 f"Добавить найденные платежи?\n\n"
                                 f"Новая сумма будет: {format_number(new_amount)}")
    
    if result:
        # Обновляем поле с новой суммой
        entry_payment.delete(0, tk.END)
        entry_payment.insert(0, str(new_amount).replace('.', ','))
        messagebox.showinfo("Успех", 
                          f"Сумма {format_number(total_found)} добавлена!\n"
                          f"Общая сумма платежей: {format_number(new_amount)}")
        
        # Если режим нескольких депозитов, обновляем доступную сумму
        if deposit_count.get() != "1" and hasattr(root, 'initial_balance'):
            update_available_balance()

def process_calculation():
    """Основная функция обработки"""
    try:
        # Проверяем существование файла 1
        if not os.path.exists(FILE1_PATH):
            messagebox.showerror("Ошибка", f"Файл не найден по пути:\n{FILE1_PATH}")
            return
        
        # 1. Читаем первый файл и находим последнее число в столбце G
        try:
            df1 = pd.read_excel(FILE1_PATH)
            if df1.empty or len(df1.columns) < 7:
                messagebox.showerror("Ошибка", "Файл 1 не содержит столбец G или пуст")
                return
            
            # Ищем последнюю непустую ячейку в столбце G (индекс 6)
            column_g = df1.iloc[:, 6].dropna()
            if column_g.empty:
                messagebox.showerror("Ошибка", "В столбце G файла 1 нет данных")
                return
            
            last_value_str = str(column_g.iloc[-1])
            initial_balance = parse_number(last_value_str)
            root.initial_balance = initial_balance  # Сохраняем для использования
            
            # Показываем найденный баланс
            label_balance.config(text=f"Найденный баланс: {format_number(initial_balance)}")
            
            # Если выбрано несколько депозитов, обновляем доступную сумму
            if deposit_count.get() != "1" and 'available_balance_label' in globals() and available_balance_label:
                update_available_balance()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка чтения файла 1: {str(e)}")
            return
        
        # Получаем сумму платежей (обязательное поле для всех вариантов)
        try:
            payment_amount = parse_number(entry_payment.get())
        except Exception as e:
            messagebox.showerror("Ошибка", "Введите корректную сумму платежей")
            return
        
        count = deposit_count.get()
        
        if count == "1":
            # Обработка одного депозита
            try:
                interest_rate = entry_rate.get().strip()
                target_date = entry_date.get().strip()
                
                if not interest_rate or not target_date:
                    messagebox.showerror("Ошибка", "Заполните все поля")
                    return
                    
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка в данных: {str(e)}")
                return
            
            # Выполняем расчеты (вычитаем платежи)
            final_amount = initial_balance - payment_amount - 150000
            rounded_amount = round(final_amount / 1000) * 1000  # Округление до тысяч
            
            # Проверяем, что сумма не отрицательная
            if rounded_amount < 0:
                messagebox.showerror("Ошибка", "Сумма к размещению не может быть отрицательной!")
                return
            
            # Форматируем сообщение
            days_info = calculate_days_until(target_date)
            
            message = f"""Добрый день!
Прошу подписать заявление на размещение депозитов в ГПБ Бизнес Онлайн

Сумма: {format_number(rounded_amount)}
Срок: до {target_date} ({days_info} дней)
Ставка: {interest_rate}%

Счёт списания и счёт зачисления 40701.810.7.00000005417

Сумма платежей сегодня: {format_number(payment_amount)}"""
            
            # Сохраняем данные для подтверждения
            root.deposits_data = [{
                'amount': rounded_amount,
                'rate': interest_rate,
                'date': target_date,
                'days': days_info
            }]
            root.total_amount = rounded_amount
            
        else:
            # Обработка нескольких депозитов
            try:
                deposits_data = []
                total_amount = 0
                
                for i, deposit in enumerate(deposit_entries):
                    amount_str = deposit['amount'].get().strip()
                    rate_str = deposit['rate'].get().strip()
                    date_str = deposit['date'].get().strip()
                    
                    if not amount_str or not rate_str or not date_str:
                        messagebox.showerror("Ошибка", f"Заполните все поля для депозита {i+1}")
                        return
                    
                    amount = parse_number(amount_str)
                    total_amount += amount
                    days_info = calculate_days_until(date_str)
                    
                    deposits_data.append({
                        'amount': amount,
                        'rate': rate_str,
                        'date': date_str,
                        'days': days_info
                    })
                    
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка в данных: {str(e)}")
                return
            
            # Доступный остаток после вычета платежей
            available_balance = initial_balance - payment_amount - 10000
            
            if total_amount > available_balance:
                messagebox.showerror("Ошибка", 
                    f"Общая сумма депозитов ({format_number(total_amount)}) превышает доступный остаток!\n\n"
                    f"Доступно: {format_number(available_balance)}\n"
                    f"Превышение: {format_number(total_amount - available_balance)}")
                return
            
            # Форматируем сообщение для нескольких депозитов
            message = "Добрый день!\n"
            message += "Прошу подписать заявление на размещение депозитов в ГПБ Бизнес Онлайн\n\n"
            
            message += f"Общая сумма: {format_number(total_amount)}\n\n"
            
            for i, deposit in enumerate(deposits_data):
                message += f"Депозит {i+1}:\n"
                message += f"Сумма: {format_number(deposit['amount'])}\n"
                message += f"Ставка: {deposit['rate']}%\n"
                message += f"Срок: до {deposit['date']} ({deposit['days']} дней)\n\n"
            
            message += f"Счёт списания и счёт зачисления 40701.810.7.00000005417\n\n"
            message += f"Сумма платежей сегодня: {format_number(payment_amount)}"
            
            # Сохраняем данные для подтверждения
            root.deposits_data = deposits_data
            root.total_amount = total_amount
        
        # Показываем результат
        result_text.delete(1.0, tk.END)
        result_text.insert(1.0, message)
        
        # Активируем кнопку подтверждения
        btn_confirm.config(state=tk.NORMAL, bg="#2196F3")
        
        # Копируем в буфер обмена
        root.clipboard_clear()
        root.clipboard_append(message)
        messagebox.showinfo("Готово", "Сообщение сформировано и скопировано в буфер обмена!\nТеперь вы можете подтвердить размещение депозитов.")
        
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
        import traceback
        traceback.print_exc()

def confirm_deposits():
    """Подтверждение и сохранение депозитов в реестр"""
    if not hasattr(root, 'deposits_data'):
        messagebox.showerror("Ошибка", "Сначала сформируйте сообщение о депозитах")
        return
    
    result = messagebox.askyesno("Подтверждение", 
                                "Вы уверены, что хотите сохранить данные о депозитах в реестр?\n\n"
                                "Это действие нельзя отменить.")
    
    if result:
        if save_to_registry(root.deposits_data, root.total_amount):
            # Деактивируем кнопку подтверждения
            btn_confirm.config(state=tk.DISABLED, bg="light gray")
        else:
            messagebox.showerror("Ошибка", "Не удалось сохранить данные в реестр")

# Создаем графический интерфейс
root = tk.Tk()
root.title("Расчет депозитов ГПБ")
root.geometry("700x750")

# СОЗДАЕМ ГЛАВНЫЙ КОНТЕЙНЕР С ПРОКРУТКОЙ
main_canvas = tk.Canvas(root)
scrollbar_vertical = tk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
scrollable_frame = tk.Frame(main_canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
)

main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
main_canvas.configure(yscrollcommand=scrollbar_vertical.set)

# Упаковываем Canvas и Scrollbar
main_canvas.pack(side="left", fill="both", expand=True)
scrollbar_vertical.pack(side="right", fill="y")

# Привязываем колесико мыши к прокрутке
def on_mousewheel(event):
    main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

main_canvas.bind_all("<MouseWheel>", on_mousewheel)

# ВСЕ ВИДЖЕТЫ ТЕПЕРЬ СОЗДАЕМ В scrollable_frame ВМЕСТО root
# Информация о файле
frame_info = tk.Frame(scrollable_frame)
frame_info.pack(pady=10, padx=20, fill=tk.X)

tk.Label(frame_info, text=f"Файл баланса: {FILE1_PATH}", 
         wraplength=600, justify=tk.LEFT, fg="blue").pack(anchor=tk.W)

label_balance = tk.Label(frame_info, text="Баланс: не загружен", fg="red")
label_balance.pack(anchor=tk.W, pady=(5, 0))

# Выбор количества депозитов
frame_deposit_count = tk.Frame(scrollable_frame)
frame_deposit_count.pack(pady=10, padx=20, fill=tk.X)

tk.Label(frame_deposit_count, text="Количество депозитов:").pack(anchor=tk.W)
deposit_count = tk.StringVar(value="1")
deposit_combo = ttk.Combobox(frame_deposit_count, 
                           textvariable=deposit_count,
                           values=["1", "2", "3", "4", "5"],
                           state="readonly",
                           width=10)
deposit_combo.pack(anchor=tk.W, pady=(5, 10))
deposit_count.trace('w', on_deposit_count_change)

# Фрейм для полей ввода
deposit_fields_frame = tk.Frame(scrollable_frame)
deposit_fields_frame.pack(pady=10, padx=20, fill=tk.X, expand=True)

# Инициализируем глобальные переменные
entry_payment = None
entry_rate = None
entry_date = None
deposit_entries = []
available_balance_label = None
remaining_label = None

# Инициализируем поля для одного депозита
on_deposit_count_change()

# Кнопка выполнения
btn_process = tk.Button(scrollable_frame, text="Сформировать сообщение", 
                       command=process_calculation,
                       bg="#4CAF50", fg="white", font=("Arial", 12, "bold"))
btn_process.pack(pady=10)

# Кнопка подтверждения
btn_confirm = tk.Button(scrollable_frame, text="Подтвердить размещение депозитов", 
                       command=confirm_deposits,
                       state=tk.DISABLED, bg="light gray", fg="black", font=("Arial", 12, "bold"))
btn_confirm.pack(pady=5)

# Поле для результата
tk.Label(scrollable_frame, text="Результат:").pack(anchor=tk.W, padx=20)
result_text = tk.Text(scrollable_frame, height=12, width=75, font=("Arial", 10))
result_text.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)

# Добавляем скроллбар для текста
scrollbar_text = tk.Scrollbar(result_text)
scrollbar_text.pack(side=tk.RIGHT, fill=tk.Y)
result_text.config(yscrollcommand=scrollbar_text.set)
scrollbar_text.config(command=result_text.yview)

# Инструкция
instruction = """
ИНСТРУКЦИЯ:
1. Выберите количество депозитов
2. Заполните сумму платежей сегодня (можно нажать кнопку "Добавить платежи от контрагентов" 
   для автоматического добавления платежей от ПОЧТА РОССИИ к текущей сумме)
3. Нажмите "Сформировать сообщение" для загрузки баланса
4. Для нескольких депозитов:
   • Используйте кнопки "Распределить поровну" или "Основной + остальные по 1000"
   • При необходимости скорректируйте суммы вручную
   • Заполните ставки и даты для каждого депозита
5. Снова нажмите "Сформировать сообщение" для финального расчета
6. Сообщение автоматически скопируется в буфер обмена
7. Нажмите "Подтвердить размещение депозитов" для сохранения в реестр

Примечание: Используйте колесико мыши для прокрутки окна при большом количестве депозитов.
"""
tk.Label(scrollable_frame, text=instruction, justify=tk.LEFT, fg="gray", 
         font=("Arial", 8)).pack(anchor=tk.W, padx=20, pady=(0, 10))

# Запускаем приложение
if __name__ == "__main__":
    root.mainloop()
