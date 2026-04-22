import streamlit as st
import subprocess
import sys
import os
import locale
import time
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook

# Настройка страницы
st.set_page_config(
    page_title="🚀 Запуск скриптов",
    page_icon="🚀",
    layout="wide"
)

# ==================== КОНФИГУРАЦИЯ СКРИПТОВ ====================
# Основная группа скриптов
MAIN_SCRIPTS_CONFIG = {
    "SFT.py": "📝 СФТ Платеж",
    "rusfar.py": "📈 Обновление Rusfar",
    "INDEX.py": "📊 Обновление индексов",
    "simple.py": "📁 Скопировать файл Daily Income 2026",
    "OSVRG.py": "🎯ОСВ",
    "check.py": "⚙️ Проверить наличие актуальных файлов в диадоке",
    "SCHA.py": "✅СЧА для ИД РЖ"
}

# Дополнительная группа скриптов
EXTRA_SCRIPTS_CONFIG = {
    "zaprosstavok.py": "📨 Запрос ставок",
    "stop.py": "❌ Убрать заглушку"
}

# ==================== 🐉 КОНФИГУРАЦИЯ ДЛЯ ПРОСМОТРА РЕКВИЗИТОВ ====================
# Путь к файлу с реквизитами компаний
COMPANY_DATA_PATH = r"C:\Users\ytggf\OneDrive\Документы\renlife\Сводные ааааа\ff\Реквизиты.xlsx"
# =============================================================================

# ==================== НАСТРОЙКИ ТАЙМАУТОВ ====================
# Можно настроить для каждого скрипта индивидуально (в секундах)
SCRIPT_TIMEOUTS = {
    # Основные скрипты
    "SFT.py": 0,
    "rusfar.py": 5,
    "INDEX.py": 5,
    "simple.py": 0,
    "OSVRG.py": 5,
    "check.py": 0,
    "SCHA.py": 40,
    
    # Дополнительные скрипты
    "zaprosstavok.py": 5,
    "stop.py": 300,
}

# Таймаут по умолчанию, если скрипт не указан в SCRIPT_TIMEOUTS
DEFAULT_TIMEOUT = 3
# ===============================================================

# Определяем системную кодировку
try:
    SYSTEM_ENCODING = locale.getpreferredencoding()
except:
    SYSTEM_ENCODING = 'cp1251'

st.title("🚀 Запуск Python скриптов")
st.markdown("---")

# Инициализация session state для хранения времени последнего нажатия
if 'last_click_time' not in st.session_state:
    st.session_state.last_click_time = {}

if 'button_cooldown' not in st.session_state:
    st.session_state.button_cooldown = {}

# ==================== 🐉 ФУНКЦИИ ДЛЯ ПРОСМОТРА РЕКВИЗИТОВ ====================
@st.cache_data(ttl=300)  # Кешируем данные на 5 минут
def load_companies_data(excel_path):
    """Загружает данные о компаниях из Excel файла"""
    if not os.path.exists(excel_path):
        return None, f"❌ Файл не найден по пути:\n{excel_path}"
    
    try:
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        
        companies = []
        max_col = sheet.max_column
        
        for col_idx in range(1, max_col + 1):
            company_name = sheet.cell(row=5, column=col_idx).value
            if company_name and str(company_name).strip():
                # Собираем данные с 6 строки
                data = []
                for row_idx in range(6, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        data.append(str(cell_value))
                
                companies.append({
                    "name": str(company_name).strip(),
                    "data": data
                })
        
        if not companies:
            return None, "❌ Не найдено компаний в 5-й строке столбцов"
        
        return companies, None
        
    except Exception as e:
        return None, f"❌ Ошибка чтения Excel:\n{str(e)}"

def show_companies_viewer():
    """Отображает интерфейс просмотра реквизитов компаний"""
    st.markdown("### 🏢 Просмотр реквизитов компаний")
    
    # Загружаем данные
    companies, error = load_companies_data(COMPANY_DATA_PATH)
    
    if error:
        st.error(error)
        st.info(f"📁 Проверьте путь к файлу:\n`{COMPANY_DATA_PATH}`")
        return
    
    if not companies:
        st.warning("Нет данных для отображения")
        return
    
    # Создаем выпадающий список компаний
    company_names = [c["name"] for c in companies]
    
    # Выбор компании
    selected_company = st.selectbox(
        "📋 Выберите компанию:",
        options=company_names,
        key="company_selector"
    )
    
    # Находим выбранную компанию
    current_company = next((c for c in companies if c["name"] == selected_company), None)
    
    if current_company:
        # Показываем информацию о компании
        col1, col2 = st.columns([3, 1])
        
        with col1:
            st.markdown(f"### 🏷️ {current_company['name']}")
        
        with col2:
            # Кнопка копирования всех данных
            if current_company["data"]:
                if st.button("📋 Копировать все данные", key="copy_all_data", use_container_width=True):
                    text_to_copy = "\n".join(current_company["data"])
                    st.write("✅ Данные скопированы в буфер обмена!")
                    st.code(text_to_copy, language="text")
                    # Для реального копирования в буфер обмена используем JS
                    st.markdown(
                        f"""
                        <script>
                        function copyToClipboard() {{
                            const text = `{text_to_copy.replace('`', '\\`')}`;
                            navigator.clipboard.writeText(text);
                        }}
                        copyToClipboard();
                        </script>
                        """,
                        unsafe_allow_html=True
                    )
        
        st.markdown("---")
        
        # Отображаем данные
        if current_company["data"]:
            st.markdown("#### 📄 Реквизиты компании:")
            
            # Создаем DataFrame для красивого отображения
            df = pd.DataFrame({
                "№": range(1, len(current_company["data"]) + 1),
                "Данные": current_company["data"]
            })
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            # Альтернативное отображение в виде текста
            with st.expander("📝 Показать в текстовом виде"):
                text_content = "\n".join([f"{i}. {item}" for i, item in enumerate(current_company["data"], 1)])
                st.code(text_content, language="text")
        else:
            st.info("⚠️ Нет данных для этой компании (ниже 5-й строки нет заполненных ячеек)")
        
        # Статистика
        st.caption(f"📊 Всего записей: {len(current_company['data'])}")
# =============================================================================

# Функция для проверки таймаута кнопки
def is_button_on_cooldown(button_key, cooldown_seconds):
    """
    Проверяет, находится ли кнопка в состоянии таймаута
    Возвращает (bool, float) - (на таймауте ли, сколько секунд осталось)
    """
    if button_key in st.session_state.last_click_time:
        last_click = st.session_state.last_click_time[button_key]
        time_diff = (datetime.now() - last_click).total_seconds()
        
        if time_diff < cooldown_seconds:
            remaining = round(cooldown_seconds - time_diff, 1)
            return True, remaining
    
    return False, 0

# Функция для обновления времени последнего нажатия
def update_last_click(button_key):
    st.session_state.last_click_time[button_key] = datetime.now()

# Функция для чтения файла инструкции
def read_instruction_file(script_name):
    """Читает содержимое txt файла с инструкцией"""
    # Заменяем расширение .py на .txt
    txt_file = script_name.replace('.py', '.txt')
    
    try:
        if os.path.exists(txt_file):
            with open(txt_file, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            return f"❌ Файл инструкции {txt_file} не найден"
    except Exception as e:
        return f"❌ Ошибка при чтении инструкции: {str(e)}"

# Функция для создания кнопок инструкции
def create_instruction_button(script_name, button_name, col):
    """Создает кнопку инструкции в указанной колонке"""
    instruction_key = f"inst_{script_name}"
    
    # Кнопка инструкции (маленькая, с эмодзи)
    if col.button("ℹ️", key=instruction_key, help=f"Инструкция для {button_name}"):
        instruction_text = read_instruction_file(script_name)
        
        # Создаем модальное окно с инструкцией
        with st.expander(f"📖 Инструкция: {button_name}", expanded=True):
            st.markdown(instruction_text)
            
            # Кнопка для закрытия
            if st.button("✖️ Закрыть", key=f"close_{instruction_key}"):
                st.rerun()

# Функция для запуска скрипта
def run_script(script_path):
    try:
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            encoding=SYSTEM_ENCODING,
            errors='replace',
            timeout=300
        )
        return {
            'success': result.returncode == 0,
            'output': result.stdout,
            'error': result.stderr
        }
    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'output': '',
            'error': f'Таймаут: скрипт выполнялся более 5 минут'
        }
    except Exception as e:
        return {
            'success': False,
            'output': '',
            'error': f'Ошибка запуска: {str(e)}'
        }

# 🐉 СОЗДАЕМ ТРИ ВКЛАДКИ (было две, добавили третью)
tab1, tab2, tab3 = st.tabs(["📋 Основные скрипты", "🔧 Дополнительные скрипты", "🏢 Реквизиты компаний"])

# === Вкладка 1: Основные скрипты ===
with tab1:
    # Проверяем какие основные скрипты существуют
    available_main_scripts = {}
    missing_main_scripts = []

    for script_file, button_name in MAIN_SCRIPTS_CONFIG.items():
        if os.path.exists(script_file):
            available_main_scripts[script_file] = button_name
        else:
            missing_main_scripts.append(script_file)

    # Показываем доступные основные скрипты
    if available_main_scripts:
        st.subheader("✅ Доступные основные скрипты:")
        
        script_items = list(available_main_scripts.items())
        
        for script_file, button_name in script_items:
            # Получаем таймаут для скрипта
            cooldown = SCRIPT_TIMEOUTS.get(script_file, DEFAULT_TIMEOUT)
            button_key = f"main_{script_file}"
            
            # Проверяем, на таймауте ли кнопка
            on_cooldown, remaining = is_button_on_cooldown(button_key, cooldown)
            
            # Создаем строку с двумя колонками: для кнопки запуска и для кнопки инструкции
            col1, col2 = st.columns([5, 1])
            
            with col1:
                # Создаем текст для кнопки
                if on_cooldown:
                    button_text = f"⏳ {button_name} (подождите {remaining}с)"
                    disabled = True
                else:
                    button_text = button_name
                    disabled = False
                
                # Кнопка запуска
                if st.button(button_text, key=button_key, use_container_width=True, disabled=disabled):
                    # Обновляем время последнего нажатия
                    update_last_click(button_key)
                    
                    with st.spinner(f"Запускаю {button_name}..."):
                        result = run_script(script_file)
                        
                        if result['success']:
                            st.success(f"✅ {button_name} выполнен!")
                            if result['output']:
                                with st.expander("📋 Показать вывод", expanded=True):
                                    st.code(result['output'])
                        else:
                            st.error(f"❌ Ошибка в {button_name}")
                            if result['error']:
                                with st.expander("🔍 Показать ошибку", expanded=True):
                                    st.code(result['error'])
            
            # Кнопка инструкции в отдельной колонке
            create_instruction_button(script_file, button_name, col2)
            
            # Добавляем небольшой отступ между строками
            st.markdown("")

        # Кнопка для запуска всех основных скриптов
        st.markdown("---")
        
        # Для кнопки "запустить все" тоже добавим таймаут
        all_button_key = "run_all_main"
        all_cooldown = 10  # 10 секунд таймаут для запуска всех скриптов
        on_cooldown_all, remaining_all = is_button_on_cooldown(all_button_key, all_cooldown)
        
        if on_cooldown_all:
            all_button_text = f"⏳ ЗАПУСТИТЬ ВСЕ ОСНОВНЫЕ СКРИПТЫ (подождите {remaining_all}с)"
            disabled_all = True
        else:
            all_button_text = "⚡ ЗАПУСТИТЬ ВСЕ ОСНОВНЫЕ СКРИПТЫ"
            disabled_all = False
        
        if st.button(all_button_text, key=all_button_key, use_container_width=True, type="secondary", disabled=disabled_all):
            update_last_click(all_button_key)
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_area = st.empty()
            
            all_output = "📊 РЕЗУЛЬТАТЫ ВЫПОЛНЕНИЯ ОСНОВНЫХ СКРИПТОВ:\n\n"
            
            for i, (script_file, button_name) in enumerate(script_items):
                progress = (i) / len(script_items)
                progress_bar.progress(progress)
                status_text.text(f"🔄 Выполняется: {button_name}")
                
                result = run_script(script_file)
                
                all_output += f"=== {button_name} ===\n"
                if result['success']:
                    all_output += "✅ УСПЕХ\n"
                    if result['output']:
                        all_output += result['output'] + "\n"
                else:
                    all_output += "❌ ОШИБКА\n"
                    if result['error']:
                        all_output += result['error'] + "\n"
                all_output += "\n"
            
            progress_bar.progress(1.0)
            status_text.text("✅ Все основные скрипты выполнены!")
            
            with results_area.container():
                st.code(all_output)

# === Вкладка 2: Дополнительные скрипты ===
with tab2:
    # Проверяем какие дополнительные скрипты существуют
    available_extra_scripts = {}
    missing_extra_scripts = []

    for script_file, button_name in EXTRA_SCRIPTS_CONFIG.items():
        if os.path.exists(script_file):
            available_extra_scripts[script_file] = button_name
        else:
            missing_extra_scripts.append(script_file)

    # Показываем доступные дополнительные скрипты
    if available_extra_scripts:
        st.subheader("✅ Доступные дополнительные скрипты:")
        
        script_items = list(available_extra_scripts.items())
        
        for script_file, button_name in script_items:
            # Получаем таймаут для скрипта
            cooldown = SCRIPT_TIMEOUTS.get(script_file, DEFAULT_TIMEOUT)
            button_key = f"extra_{script_file}"
            
            # Проверяем, на таймауте ли кнопка
            on_cooldown, remaining = is_button_on_cooldown(button_key, cooldown)
            
            # Создаем строку с двумя колонками: для кнопки запуска и для кнопки инструкции
            col1, col2 = st.columns([5, 1])
            
            with col1:
                # Создаем текст для кнопки
                if on_cooldown:
                    button_text = f"⏳ {button_name} (подождите {remaining}с)"
                    disabled = True
                else:
                    button_text = button_name
                    disabled = False
                
                # Кнопка запуска
                if st.button(button_text, key=button_key, use_container_width=True, disabled=disabled):
                    # Обновляем время последнего нажатия
                    update_last_click(button_key)
                    
                    with st.spinner(f"Запускаю {button_name}..."):
                        result = run_script(script_file)
                        
                        if result['success']:
                            st.success(f"✅ {button_name} выполнен!")
                            if result['output']:
                                with st.expander("📋 Показать вывод", expanded=True):
                                    st.code(result['output'])
                        else:
                            st.error(f"❌ Ошибка в {button_name}")
                            if result['error']:
                                with st.expander("🔍 Показать ошибку", expanded=True):
                                    st.code(result['error'])
            
            # Кнопка инструкции в отдельной колонке
            create_instruction_button(script_file, button_name, col2)
            
            # Добавляем небольшой отступ между строками
            st.markdown("")

        # Кнопка для запуска всех дополнительных скриптов
        st.markdown("---")
        
        # Для кнопки "запустить все" тоже добавим таймаут
        all_button_key = "run_all_extra"
        all_cooldown = 10  # 10 секунд таймаут для запуска всех скриптов
        on_cooldown_all, remaining_all = is_button_on_cooldown(all_button_key, all_cooldown)
        
        if on_cooldown_all:
            all_button_text = f"⏳ ЗАПУСТИТЬ ВСЕ ДОПОЛНИТЕЛЬНЫЕ СКРИПТЫ (подождите {remaining_all}с)"
            disabled_all = True
        else:
            all_button_text = "⚡ ЗАПУСТИТЬ ВСЕ ДОПОЛНИТЕЛЬНЫЕ СКРИПТЫ"
            disabled_all = False
        
        if st.button(all_button_text, key=all_button_key, use_container_width=True, type="secondary", disabled=disabled_all):
            update_last_click(all_button_key)
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            results_area = st.empty()
            
            all_output = "📊 РЕЗУЛЬТАТЫ ВЫПОЛНЕНИЯ ДОПОЛНИТЕЛЬНЫХ СКРИПТОВ:\n\n"
            
            for i, (script_file, button_name) in enumerate(script_items):
                progress = (i) / len(script_items)
                progress_bar.progress(progress)
                status_text.text(f"🔄 Выполняется: {button_name}")
                
                result = run_script(script_file)
                
                all_output += f"=== {button_name} ===\n"
                if result['success']:
                    all_output += "✅ УСПЕХ\n"
                    if result['output']:
                        all_output += result['output'] + "\n"
                else:
                    all_output += "❌ ОШИБКА\n"
                    if result['error']:
                        all_output += result['error'] + "\n"
                all_output += "\n"
            
            progress_bar.progress(1.0)
            status_text.text("✅ Все дополнительные скрипты выполнены!")
            
            with results_area.container():
                st.code(all_output)

# 🐉 === Вкладка 3: Реквизиты компаний (НОВАЯ) ===
with tab3:
    show_companies_viewer()

# Инструкция для пользователей
with st.sidebar:
    st.header("📖 Инструкция")
    
    st.markdown("""
    Здесь будет ссылка на confluence
    """)
    
    st.markdown("---")
    st.caption("🔄 Для обновления списка скриптов перезапусти приложение")









