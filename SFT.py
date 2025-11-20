import os
import datetime
import openpyxl
from pathlib import Path

def analyze_excel_data():
    # Путь 1 - основная папка с месяцами
    base_path = r"Q:\Операционный отдел\!!! ВЫПЛАТЫ НПФ (Платеж СФТ)"  # Замените на ваш путь
    
    # Путь 2 - файл для записи результата
    result_file_path = r"M:\Финансовый департамент\Treasury\Базы данных(автоматизация)\DI_DATABASE\Данные.xlsx"  # Замените на ваш путь
    
    try:
        # Получаем текущую дату
        today = datetime.date.today()
        current_month = today.month
        current_date_str = today.strftime("%d.%m.%Y")
        
        print(f"Текущая дата: {current_date_str}")
        print(f"Текущий месяц: {current_month}")
        
        # Ищем папку с текущим месяцем
        month_folders = os.listdir(base_path)
        target_month_folder = None
        
        for folder in month_folders:
            if folder.startswith(f"{current_month:02d}-"):
                target_month_folder = folder
                break
        
        if not target_month_folder:
            print(f"Папка для месяца {current_month} не найдена")
            return
        
        print(f"Найдена папка месяца: {target_month_folder}")
        
        # Путь к папке месяца
        month_path = os.path.join(base_path, target_month_folder)
        
        # Ищем папку с сегодняшней датой и пометкой ГОТОВО
        date_folders = os.listdir(month_path)
        target_date_folder = None
        
        for folder in date_folders:
            if current_date_str in folder and "ГОТОВО" in folder:
                target_date_folder = folder
                break
        
        if not target_date_folder:
            print(f"Папка с датой {current_date_str} и пометкой 'ГОТОВО' не найдена")
            # Проверим, есть ли папка с датой, но без пометки ГОТОВО
            for folder in date_folders:
                if current_date_str in folder and "ГОТОВО" not in folder:
                    print("Найдена папка с датой, но без пометки 'ГОТОВО' - файл не готов")
                    return
            return
        
        print(f"Найдена папка с датой: {target_date_folder}")
        
        # Путь к папке с датой
        date_folder_path = os.path.join(month_path, target_date_folder)
        
        # Ищем Excel файл с сегодняшней датой
        excel_files = [f for f in os.listdir(date_folder_path) 
                      if f.endswith(('.xlsx', '.xls')) and current_date_str in f]
        
        if not excel_files:
            print(f"Excel файл с датой {current_date_str} не найден")
            return
        
        target_excel_file = excel_files[0]
        print(f"Найден Excel файл: {target_excel_file}")
        
        # Открываем Excel файл
        excel_path = os.path.join(date_folder_path, target_excel_file)
        workbook = openpyxl.load_workbook(excel_path, data_only=True)  # data_only=True для получения значений, а не формул
        sheet = workbook.active
        
        # Ищем столбец "Сумма реквеста" и последнее значение
        amount_column = None
        last_amount = None
        
        # Ищем заголовок "Сумма реквеста"
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value and "Сумма реквеста" in str(cell_value):
                amount_column = col
                break
        
        if not amount_column:
            print("Столбец 'Сумма реквеста' не найден")
            return
        
        print(f"Столбец 'Сумма реквеста' найден в колонке {amount_column}")
        
        # Ищем последнее значение в столбце (игнорируя заголовок и пустые ячейки)
        for row in range(sheet.max_row, 1, -1):
            cell_value = sheet.cell(row=row, column=amount_column).value
            if cell_value is not None and cell_value != "":
                # Преобразуем значение в число, если возможно
                try:
                    if isinstance(cell_value, (int, float)):
                        last_amount = cell_value
                    else:
                        # Пытаемся преобразовать строку в число
                        cleaned_value = str(cell_value).replace(',', '.').replace(' ', '')
                        last_amount = float(cleaned_value)
                    break
                except ValueError:
                    print(f"Не удалось преобразовать значение '{cell_value}' в число")
                    continue
        
        if last_amount is None:
            print("Не найдены числовые данные в столбце 'Сумма реквеста'")
            return
        
        print(f"Найдено последнее числовое значение: {last_amount}")
        
        # Записываем значение в файл по пути 2
        if os.path.exists(result_file_path):
            result_workbook = openpyxl.load_workbook(result_file_path)
        else:
            result_workbook = openpyxl.Workbook()
        
        result_sheet = result_workbook.active
        result_sheet['B87'] = last_amount
        result_workbook.save(result_file_path)
        
        print(f"Значение {last_amount} успешно записано в файл {result_file_path} в ячейку B87")
        
    except Exception as e:
        print(f"Произошла ошибка: {e}")

if __name__ == "__main__":
    analyze_excel_data()
